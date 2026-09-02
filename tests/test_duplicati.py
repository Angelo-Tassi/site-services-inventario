"""Doppioni: dentro un foglio importato e dentro l'inventario.

Il programma da solo non crea duplicati - l'inserimento li rifiuta e
l'importazione aggiorna invece di duplicare - ma il file dati e' un .xlsx che si
puo' aprire e correggere a mano, ed e' da li' che entrano.
"""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import Workbook, load_workbook
from inventario.store import (InventoryError, InventoryStore, new_item,
                              rows_from_workbook)

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
STANZE = [BAU, KIOSK, DR]

# ---- un foglio che contiene due volte lo stesso identificativo lo dice
wb = Workbook(); ws = wb.active; ws.title = "Inventario"
ws.append(["Asset Tag", "Tipo", "Stanza", "Note"])
ws.append(["IT-7001", "Laptop", BAU, "prima riga"])
ws.append(["IT-7002", "Laptop", BAU, ""])
ws.append(["IT-7001", "Laptop", KIOSK, "seconda, vince questa"])
foglio = os.path.join(tempfile.mkdtemp(), "doppi.xlsx")
wb.save(foglio); wb.close()

items, esito = rows_from_workbook(foglio, STANZE)
assert esito["doppioni"] == ["IT-7001"], esito["doppioni"]
assert len(items) == 3, len(items)

# ---- importandolo, vale l'ultima riga e in inventario ce n'e' una sola
percorso = fixture.build()
store = InventoryStore(percorso, iphone_room=BAU)
store.load()
store.import_items(items, "merge")
store.load()
uno = [i for i in store.items if i["asset_tag"] == "IT-7001"]
assert len(uno) == 1, uno
assert uno[0]["stanza"] == KIOSK, uno[0]

# ---- l'inserimento singolo rifiuta un identificativo gia' presente
try:
    store.add(new_item("IT-7001", "Laptop", "T14", "PF1", DR, ""))
    raise AssertionError("doveva rifiutare il doppione")
except InventoryError as exc:
    assert "IT-7001" in str(exc) and "gia' in inventario" in str(exc), exc
    assert "Non e' stato inserito niente" in str(exc), exc
store.load()
assert len([i for i in store.items if i["asset_tag"] == "IT-7001"]) == 1

# ---- doppioni entrati modificando il file a mano: si trovano e si tolgono
wb = load_workbook(percorso); ws = wb.active
righe = list(ws.iter_rows(values_only=True))
ws.append(list(righe[1]))
ws.append(list(righe[1]))          # la stessa riga tre volte in tutto
ws.append(list(righe[2]))
wb.save(percorso); wb.close()

store = InventoryStore(percorso, iphone_room=BAU)
gruppi, seriali = store.trova_duplicati()
assert len(gruppi) == 2, [(t, len(e)) for t, e in gruppi]
assert sum(len(e) - 1 for _t, e in gruppi) == 3, gruppi

prima = len(store.load())
rapporto = store.rimuovi_duplicati()
assert len(rapporto["eliminati"]) == 3, rapporto["eliminati"]
assert rapporto["prima"] == prima and rapporto["dopo"] == prima - 3
assert rapporto["copia"] and os.path.exists(rapporto["copia"])
assert store.trova_duplicati()[0] == [], "non ne devono restare"

# ---- di ogni gruppo resta la registrazione modificata piu' di recente
tenuti = {i["asset_tag"]: i for i in rapporto["tenuti"]}
for tag, tenuto in tenuti.items():
    rimasto = [i for i in store.items if i["asset_tag"] == tag]
    assert len(rimasto) == 1, rimasto
    assert rimasto[0]["modificato_il"] == tenuto["modificato_il"]

# ---- un numero di serie ripetuto su due dispositivi diversi si segnala, e basta
store.add(new_item("IT-8001", "Laptop", "T14", "SERIALE-X", DR, ""))
store.add(new_item("IT-8002", "Laptop", "T14", "SERIALE-X", DR, ""))
gruppi, seriali = store.trova_duplicati()
assert gruppi == [], gruppi
assert len(seriali) == 1 and seriali[0][0] == "SERIALE-X", seriali
rapporto = store.rimuovi_duplicati()
assert rapporto["eliminati"] == [], "un seriale ripetuto non si cancella da solo"
assert len([i for i in store.load() if i["seriale"] == "SERIALE-X"]) == 2

# ---- la chiave e' l'asset tag, non il numero di serie: due dispositivi
# diversi con lo stesso seriale restano due dispositivi
pulito2 = InventoryStore(fixture.build(), iphone_room=BAU)
pulito2.load()
pulito2.add(new_item("IT-9001", "Laptop", "T14", "STESSO-SERIALE", DR, ""))
pulito2.add(new_item("IT-9002", "Laptop", "T14", "STESSO-SERIALE", DR, ""))
gruppi, seriali = pulito2.trova_duplicati()
assert gruppi == [], "seriali uguali non fanno un duplicato"
assert len(seriali) == 1, seriali
prima = len(pulito2.load())
assert pulito2.rimuovi_duplicati()["eliminati"] == []
assert len(pulito2.load()) == prima, "non si elimina niente per un seriale ripetuto"

# ---- mentre due righe con lo stesso asset tag e seriali diversi sono duplicati
pulito3 = InventoryStore(fixture.build(), iphone_room=BAU)
pulito3.load()
wb = load_workbook(pulito3.path); ws = wb.active
righe_file = list(ws.iter_rows(values_only=True))
riga = list(righe_file[1])
intestazioni = [str(c or "") for c in righe_file[0]]
riga[intestazioni.index("Numero di serie")] = "SERIALE-DIVERSO"
ws.append(riga)
wb.save(pulito3.path); wb.close()
gruppi, _seriali = pulito3.trova_duplicati()
assert len(gruppi) == 1, gruppi
assert len(pulito3.rimuovi_duplicati()["eliminati"]) == 1

# ---- un inventario pulito non ha niente da segnalare
pulito = InventoryStore(fixture.build(), iphone_room=BAU)
assert pulito.trova_duplicati() == ([], [])
assert pulito.rimuovi_duplicati()["eliminati"] == []

print("DUPLICATI OK")
