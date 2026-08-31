"""Archivio dati: CRUD, prestiti, stati, import/export, accessi concorrenti."""
import os, sys, tempfile, threading
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import Workbook, load_workbook
from inventario import lingua as lang
lang.imposta(lang.ITALIANO)
from inventario import excel_io
from inventario.store import (ALL_FIELDS, DISPONIBILE, HEADERS, InventoryError,
                              InventoryStore, NON_DISPONIBILE, new_item,
                              rows_from_workbook)

BAU, KIOSK = "Site Services BAU", "Digital Kiosk"
d = tempfile.mkdtemp()
p = os.path.join(d, "Inventario.xlsx")
s = InventoryStore(p, iphone_room=BAU)
assert s.create_if_missing() is True
s.load(); assert s.items == []

s.add(new_item(" it-0101 ", "Laptop", "Lenovo ThinkPad T14 Gen 4", "PF4A1B2C", BAU, "reception"))
s.add(new_item("IT-0107", "Laptop", "Lenovo ThinkPad T14 Gen 5", "PF5K9M8F", KIOSK))
s.load()
assert sorted(i["asset_tag"] for i in s.items) == ["IT-0101", "IT-0107"]
assert all(i["stato"] == DISPONIBILE for i in s.items)

try:
    s.add(new_item("it-0101", "Laptop", "X", "", BAU)); raise SystemExit("duplicato accettato")
except InventoryError:
    pass

# ---------------------------------------------------------------- prestiti
quando = s.lend("IT-0107", "  Marco Bianchi ")
s.load()
loaned = [i for i in s.items if i["asset_tag"] == "IT-0107"][0]
assert loaned["prestato_a"] == "Marco Bianchi" and loaned["prestato_il"] == quando
assert loaned["stato"] == NON_DISPONIBILE
for prova in (lambda: s.lend("IT-0107", "Altro"), lambda: s.lend("IT-0101", "   ")):
    try:
        prova(); raise SystemExit("prestito non valido accettato")
    except InventoryError:
        pass

edited = dict(loaned); edited["modello"] = "T14 Gen 5 rev."
s.update("IT-0107", edited); s.load()
loaned = [i for i in s.items if i["asset_tag"] == "IT-0107"][0]
assert loaned["prestato_a"] == "Marco Bianchi", "la modifica non deve azzerare il prestito"

assert s.give_back("IT-0107") == "Marco Bianchi"
s.load()
assert [i for i in s.items if i["asset_tag"] == "IT-0107"][0]["stato"] == DISPONIBILE
try:
    s.give_back("IT-0107"); raise SystemExit("rientro doppio accettato")
except InventoryError:
    pass

# ------------------------------------------------------------- note e stati
assert s.set_note("IT-0101", "nuova nota  al   volo") is True
s.load()
assert [i for i in s.items if i["asset_tag"] == "IT-0101"][0]["note"] == "nuova nota al volo"
assert s.set_note("IT-0101", "nuova nota al volo") is False
assert s.set_stato("IT-0101", "Da rebuildare") is True
assert s.set_stato("IT-0101", "Da rebuildare") is False
try:
    s.set_stato("IT-0101", "Fantasia"); raise SystemExit("stato inventato accettato")
except InventoryError:
    pass
s.load()
assert [i for i in s.items if i["asset_tag"] == "IT-0101"][0]["stato"] == "Da rebuildare"

# ------- l'export contiene l'inventario, non la cronaca di chi l'ha toccato
exp = os.path.join(d, "export.xlsx")
excel_io.export(s.items, exp)
wb = load_workbook(exp)
assert [c.value for c in wb.active[1]] == [HEADERS[f] for f in excel_io.CAMPI_ESPORTAZIONE]
wb.close()

# ------------- l'inventario si crea da zero importando un file Excel
foreign = os.path.join(d, "da_importare.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["Asset", "Tipologia", "Model", "S/N", "Ubicazione", "IMEI",
           "In prestito a", "Prestato il", "Commenti"])
ws.append(["it-0301", "Laptop", "T14 Gen 4", "PF4Z9Q1A", KIOSK, "",
           "Elena Rossi", "29/08/2026 09:12", "kiosk 3"])
ws.append(["", "Iphone", "iPhone 13", "", KIOSK, "351111111111111", "", "", ""])
ws.append([None] * 9)
ws.append(["", "Laptop", "senza identificativo", "", KIOSK, "", "", "", ""])
wb.save(foreign); wb.close()

items, esito = rows_from_workbook(foreign)
scartate = esito['scartate']
assert len(items) == 1 and scartate == 1, (len(items), scartate)
assert esito["iphone"] == 1, "gli iPhone non si importano"

vuoto = InventoryStore(os.path.join(d, "Nuovo.xlsx"), iphone_room=BAU)
vuoto.create_if_missing()
# un iPhone gia' in inventario non deve sparire con una sostituzione
vuoto.add(new_item(tipo="Iphone", modello="Apple iPhone 14",
                   imei="356938035643809", restituito_da="M. B."))
e = vuoto.import_items(items, "replace")
assert (e["aggiunti"], e["aggiornati"]) == (1, 0), e
vuoto.load()
assert len(vuoto.items) == 2, [i["asset_tag"] for i in vuoto.items]
tel = [i for i in vuoto.items if i["asset_tag"] == "356938035643809"][0]
assert tel["stanza"] == BAU and tel["stato"] == "Da Rispedire", tel
prestato = [i for i in vuoto.items if i["asset_tag"] == "IT-0301"][0]
assert prestato["stato"] == NON_DISPONIBILE and prestato["prestato_a"] == "Elena Rossi"

# Round trip: quello che l'esportazione porta via, l'importazione lo rimette.
# Gli iPhone restano fuori dall'esportazione: si confrontano solo gli altri.
rt = os.path.join(d, "roundtrip.xlsx")
excel_io.export(vuoto.items, rt)
back, _ = rows_from_workbook(rt)
attesi = [i for i in vuoto.items if i["tipo"].lower() != "iphone"]
assert len(back) == len(attesi) == 1, (len(back), len(attesi))
for a, b in zip(sorted(attesi, key=lambda i: i["asset_tag"]),
                sorted(back, key=lambda i: i["asset_tag"])):
    for f in excel_io.CAMPI_ESPORTAZIONE:
        if f == "stato":
            continue          # dipende dal prestito, che nel file non c'e'
        assert a[f] == b[f], (f, a[f], b[f])

# Un prestito in corso NON sopravvive a un giro export/import: e' un dato di
# consultazione, sta nell'inventario e non nei file che ne escono. Chi deve
# ricostruire l'inventario per intero non usa un export ma una copia del file,
# che invece si porta dietro tutto.
assert not back[0]["prestato_a"], back[0]
assert back[0]["stato"] == DISPONIBILE, back[0]
copia = os.path.join(d, "copia_intera.xlsx")
vuoto.copia_in(copia)
identici = InventoryStore(copia).load()
prestato_nella_copia = [i for i in identici if i["asset_tag"] == "IT-0301"][0]
assert prestato_nella_copia["prestato_a"] == "Elena Rossi", prestato_nella_copia
assert prestato_nella_copia["stato"] == NON_DISPONIBILE

# l'esportazione non contiene mai iPhone, la stampa interna si'
solo_export, _ = rows_from_workbook(rt)
assert not any(i["tipo"].lower() == "iphone" for i in solo_export)
stampa_iphone = excel_io.build_print_file(vuoto.items, rooms=[BAU])
assert os.path.getsize(stampa_iphone) > 1000

# stampa per stanza
stampa = excel_io.build_print_file(vuoto.items, group_by_room=True, rooms=[BAU, KIOSK])
assert os.path.getsize(stampa) > 1000

# ------------------------------------------------------------- concorrenza
errori = []
def worker(n):
    st = InventoryStore(p, iphone_room=BAU)
    for k in range(5):
        try:
            st.add(new_item("T%d-%02d" % (n, k), "Laptop", "T14", "SN%d%d" % (n, k), KIOSK))
        except Exception as exc:
            errori.append(exc)
ts = [threading.Thread(target=worker, args=(n,)) for n in range(8)]
[t.start() for t in ts]; [t.join() for t in ts]
s.load()
assert not errori, errori
assert len(s.items) == 42, len(s.items)
assert len(set(i["asset_tag"] for i in s.items)) == 42
print("STORE OK")
