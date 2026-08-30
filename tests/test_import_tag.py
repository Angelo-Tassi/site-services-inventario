"""Importazione di un inventario unico diviso per stanza con righe-separatore."""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import Workbook
from inventario.store import (InventoryStore, riga_tag, rows_from_workbook,
                              tag_stanze)

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
STANZE = [BAU, KIOSK, DR]

# ---- i tag si ricavano dai nomi delle stanze
tags = tag_stanze(STANZE)
assert tags["BAU"] == BAU and tags["KIOSK"] == KIOSK and tags["DISASTER"] == DR
assert tags["SITE SERVICES BAU"] == BAU
assert tags["MAGAZZINO DISASTER RECOVERY"] == DR
assert "IL" not in tags and "DI" not in tags

# ---- riconoscimento della riga-separatore
assert riga_tag(["KIOSK"], tags) == KIOSK
assert riga_tag(["  bau  "], tags) == BAU
assert riga_tag(["DISASTER:"], tags) == DR
assert riga_tag([None, "kiosk", None, ""], tags) == KIOSK
assert riga_tag(["KIOSK", "IT-0101"], tags) is None, "due celle scritte: non e' un separatore"
assert riga_tag(["IT-0101"], tags) is None
assert riga_tag([], tags) is None

# ---- un unico foglio diviso dai tag
d = tempfile.mkdtemp()
foglio = os.path.join(d, "inventario_completo.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["Asset Tag", "Tipo", "Modello", "Numero di serie", "IMEI", "Note"])
ws.append(["BAU", None, None, None, None, None])
ws.append(["IT-0101", "Laptop", "Lenovo ThinkPad T14 Gen 4", "PF4A1B2C", "", "reception"])
ws.append(["IT-0104", "Tablet", "Dell Latitude 7320 Detachable", "8H2KLM3", "", ""])
ws.append([None] * 6)
ws.append([None, "KIOSK", None, None, None, None])
ws.append(["IT-0106", "Laptop", "Lenovo ThinkPad T14 Gen 5", "PF5K9M8F", "", "kiosk 1"])
ws.append(["IT-0107", "Laptop", "Lenovo ThinkPad T14 Gen 4", "PF4A2C1G", "", ""])
ws.append(["disaster", None, None, None, None, None])
ws.append(["DR-0201", "Laptop", "Lenovo ThinkPad T14 Gen 4", "PF4B7T1J", "", "scorta"])
ws.append(["", "Iphone", "Apple iPhone 14", "", "356938035643809", "restituito"])
ws.append(["", "Laptop", "senza identificativo", "", "", ""])
wb.save(foglio); wb.close()

items, esito = rows_from_workbook(foglio, STANZE)
scartate, da_tag = esito['scartate'], esito['da_tag']
assert scartate == 1, scartate
assert da_tag == 5, da_tag
assert esito["iphone"] == 1, esito      # gli iPhone non si importano
per_stanza = {}
for i in items:
    per_stanza.setdefault(i["stanza"], []).append(i["asset_tag"])
assert sorted(per_stanza[BAU]) == ["IT-0101", "IT-0104"], per_stanza
assert sorted(per_stanza[KIOSK]) == ["IT-0106", "IT-0107"], per_stanza
assert sorted(per_stanza[DR]) == ["DR-0201"], per_stanza
assert not any(i["tipo"].lower() == "iphone" for i in items)

# ---- caricamento in un inventario vuoto
p = os.path.join(d, "Inventario.xlsx")
s = InventoryStore(p, iphone_room=BAU)
s.create_if_missing()
e = s.import_items(items, "replace")
assert (e["aggiunti"], e["aggiornati"]) == (5, 0), e
s.load()
conteggi = {}
for i in s.items:
    conteggi[i["stanza"]] = conteggi.get(i["stanza"], 0) + 1
assert conteggi == {BAU: 2, KIOSK: 2, DR: 1}, conteggi

# ---- senza separatori si comporta come prima
semplice = os.path.join(d, "semplice.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["Asset Tag", "Tipo", "Modello", "Stanza"])
ws.append(["IT-0900", "Laptop", "T14", KIOSK])
wb.save(semplice); wb.close()
items, esito = rows_from_workbook(semplice, STANZE)
da_tag = esito['da_tag']
assert da_tag == 0 and items[0]["stanza"] == KIOSK

# ---- il separatore ha la precedenza sulla colonna Stanza
misto = os.path.join(d, "misto.xlsx")
wb = Workbook(); ws = wb.active
ws.append(["Asset Tag", "Tipo", "Modello", "Stanza"])
ws.append(["DISASTER", None, None, None])
ws.append(["IT-0901", "Laptop", "T14", KIOSK])
wb.save(misto); wb.close()
items, esito = rows_from_workbook(misto, STANZE)
da_tag = esito['da_tag']
assert da_tag == 1 and items[0]["stanza"] == DR, items
print("IMPORT CON TAG OK")
