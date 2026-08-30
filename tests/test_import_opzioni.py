"""Scelta iniziale dell'importazione: ambito, modalita' e conferme."""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from tkinter import messagebox
from inventario import excel_io
from inventario.store import new_item, rows_from_workbook
from inventario.ui import (PAROLA_RESET, App, ImportDialog, ImportOptionsDialog,
                           nome_file)

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
app = App(fixture.build()); app._initial_load()
TIPO = app.iphone_type()
avvisi = []
messagebox.showwarning = lambda t, m, **k: avvisi.append((t, m))
app._run(lambda: app.store.add(new_item(tipo=TIPO, modello="Apple iPhone 14",
                                        imei="356938035643809", restituito_da="M. B.")), "ok")
d = tempfile.mkdtemp()

# ------------------------------------------------ il primo popup
dlg = ImportOptionsDialog(app, app.cfg["rooms"])
assert dlg.var_ambito.get() == "tutto" and dlg.var_mode.get() == "merge"
assert str(dlg.combo.cget("state")) == "disabled", "la stanza si sceglie solo se serve"
dlg.var_ambito.set("stanza"); dlg._aggiorna()
assert str(dlg.combo.cget("state")) == "readonly"
dlg.var_stanza.set(KIOSK); dlg.var_mode.set("replace"); dlg._ok()
assert dlg.result == {"stanza": KIOSK, "mode": "replace"}, dlg.result

# ------------------------------------------------ quanti verrebbero eliminati
assert app.contati_in_eliminazione({"stanza": None, "mode": "merge"}) == 0
assert app.contati_in_eliminazione({"stanza": KIOSK, "mode": "replace"}) == 5
tutti = app.contati_in_eliminazione({"stanza": None, "mode": "replace"})
assert tutti == 13, tutti          # i 13 non-iPhone: il telefono non si conta

# ------------------------------------------------ la conferma finale
percorso = os.path.join(d, "file.xlsx")
excel_io.export([new_item("IT-0900", "Laptop", "T14 Gen 5", "PFNEW1", "")], percorso)

# unione: nessuna frase da scrivere
conf = ImportDialog(app, percorso, 1, {}, {"stanza": None, "mode": "merge"}, 0)
assert conf.sostituzione_totale is False
conf._ok(); assert conf.result == {"stanza": None, "mode": "merge"}

# sostituzione di una stanza: avviso, ma niente frase
conf = ImportDialog(app, percorso, 1, {}, {"stanza": KIOSK, "mode": "replace"}, 5)
assert conf.sostituzione_totale is False
conf._ok(); assert conf.result["stanza"] == KIOSK

# sostituzione totale: serve la frase esatta
conf = ImportDialog(app, percorso, 1, {}, {"stanza": None, "mode": "replace"}, 13)
assert conf.sostituzione_totale is True
avvisi.clear(); conf._ok()
assert conf.result is None and avvisi[-1][0] == "Conferma non valida"
assert conf.winfo_exists()
conf.var_conferma.set("elimina tutto")
conf._ok()
assert conf.result == {"stanza": None, "mode": "replace"}

# ------------------------------------------------ importazione di una sola stanza
sorgente = os.path.join(d, "solo_kiosk.xlsx")
excel_io.export([new_item("IT-0801", "Laptop", "T14 Gen 5", "PF801", DR),
                 new_item("IT-0802", "Tablet", "Dell Latitude", "8H802", BAU)], sorgente)
items, _ = rows_from_workbook(sorgente, app.cfg["rooms"])
prima_bau = sum(1 for i in app.store.items if i["stanza"] == BAU)
risultato = app._run(lambda: app.store.import_items(items, "replace", KIOSK), "ok")
assert risultato["aggiunti"] == 2 and risultato["eliminati"] == 5, risultato
assert risultato["copia"] and os.path.exists(risultato["copia"])
app.store.load()
# le righe finiscono nella stanza scelta, non in quella scritta nel file
assert app._item_by_tag("IT-0801")["stanza"] == KIOSK
assert app._item_by_tag("IT-0802")["stanza"] == KIOSK
assert sum(1 for i in app.store.items if i["stanza"] == KIOSK) == 2
# le altre stanze non si toccano, iPhone compreso
assert sum(1 for i in app.store.items if i["stanza"] == BAU) == prima_bau
assert app._item_by_tag("356938035643809") is not None

# ------------------------------------------------ sostituzione totale
risultato = app._run(lambda: app.store.import_items(items, "replace", None), "ok")
app.store.load()
assert risultato["eliminati"] == 10, risultato   # tutti i non-iPhone rimasti
tag = sorted(i["asset_tag"] for i in app.store.items)
assert tag == ["356938035643809", "IT-0801", "IT-0802"], tag

# ------------------------------------------------ il file di stanza si dichiara
uscita = os.path.join(d, "esporta_kiosk.xlsx")
excel_io.export([i for i in app.store.items if i["stanza"] == KIOSK], uscita,
                rooms=[KIOSK], titolo=KIOSK)
wb = load_workbook(uscita); ws = wb.active
assert ws.title == KIOSK, ws.title
assert ws["A1"].value == KIOSK, ws["A1"].value
assert "Esportato il" in ws["A2"].value
intestazioni = [c.value for c in ws[4]]
assert intestazioni[0] == "Asset Tag" and "Stanza" in intestazioni
wb.close()
assert nome_file(KIOSK) in "Inventario_%s_20260830.xlsx" % nome_file(KIOSK)
app.destroy()
print("OPZIONI IMPORT OK")
