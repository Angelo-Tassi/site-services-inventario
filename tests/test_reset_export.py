"""Svuotamento dell'inventario ed esportazione della singola stanza."""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from inventario import excel_io
from inventario.store import new_item, puo_essere_eliminato, rows_from_workbook
from inventario.ui import PAROLA_RESET, App, ResetDialog, nome_file

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
app = App(fixture.build()); app._initial_load()
TIPO = app.iphone_type()

# ------------------------------------------------ esportazione di una stanza
app.show_room(KIOSK)
attesi = [i["asset_tag"] for i in app.store.items if i["stanza"] == KIOSK]
assert len(attesi) == 5
d = tempfile.mkdtemp()
percorso = os.path.join(d, "solo_kiosk.xlsx")
excel_io.export([i for i in app.store.items if i["stanza"] == KIOSK], percorso, rooms=[KIOSK])
usciti, _ = rows_from_workbook(percorso, app.cfg["rooms"])
assert sorted(i["asset_tag"] for i in usciti) == sorted(attesi), usciti
assert all(i["stanza"] == KIOSK for i in usciti)
# il nome del file e' utilizzabile
assert nome_file(KIOSK) == "Digital_Kiosk"
assert nome_file("Magazzino Disaster Recovery") == "Magazzino_Disaster_Recovery"
assert "/" not in nome_file("Stanza / 2")
# l'esportazione di stanza ignora i filtri della vista
app.var_search.set("IT-0107"); app.refresh_table()
assert len(app.visible) == 1
da_esportare = [i for i in app.store.items if i["stanza"] == KIOSK]
assert len(da_esportare) == 5, "esporta la stanza, non la vista filtrata"
app.reset_filters()

# ------------------------------------------------ la conferma del reset
app.show_home()
protetti_prima = [i for i in app.store.items if not puo_essere_eliminato(i)[0]]
assert protetti_prima == []
dlg = ResetDialog(app, 13, 0)
assert dlg.parola_giusta() is False
dlg.var_conferma.set("si")
assert dlg.parola_giusta() is False
dlg._ok()
assert dlg.result is None, "senza la frase esatta non si procede"
assert dlg.winfo_exists()
dlg.var_conferma.set("  elimina tutto  ")
assert dlg.parola_giusta() is True, "maiuscole e spazi non contano"
dlg.var_conferma.set(PAROLA_RESET)
dlg._ok()
assert dlg.result is True and not dlg.winfo_exists()

# ------------------------------------------------ il reset vero
app._run(lambda: app.store.add(new_item(tipo=TIPO, modello="Apple iPhone 14",
                                        imei="356938035643809", restituito_da="M. B.")), "ok")
app._run(lambda: app.store.add(new_item(tipo=TIPO, modello="Apple iPhone 13",
                                        imei="351234567890123", restituito_da="E. R.")), "ok")
app._run(lambda: app.store.ship("351234567890123"), "ok")
prima = len(app.store.items)
assert prima == 15
protetti = [i["asset_tag"] for i in app.store.items if not puo_essere_eliminato(i)[0]]
assert sorted(protetti) == ["351234567890123", "356938035643809"], protetti

eliminati, tenuti, copia = app._run(lambda: app.store.reset(), "ok")
assert (eliminati, tenuti) == (13, 2), (eliminati, tenuti)
app.store.load()
assert sorted(i["asset_tag"] for i in app.store.items) == sorted(protetti)

# la copia di sicurezza contiene tutto quello che c'era prima
assert os.path.exists(copia), copia
assert "prima_del_reset" in os.path.basename(copia)
assert os.path.dirname(copia) == os.path.dirname(app.store.path)
wb = load_workbook(copia)
righe = sum(1 for r in wb.active.iter_rows(min_row=2, values_only=True) if any(r))
wb.close()
assert righe == prima, (righe, prima)

# ------------------------------------------------ si ricarica da un'importazione
sorgente = os.path.join(d, "reimport.xlsx")
excel_io.export([new_item("IT-0101", "Laptop", "T14 Gen 4", "PF4A1B2C", BAU),
                 new_item("IT-0106", "Laptop", "T14 Gen 5", "PF5K9M8F", KIOSK)], sorgente)
items, _ = rows_from_workbook(sorgente, app.cfg["rooms"])
assert app._run(lambda: app.store.import_items(items, "replace"), "ok") == (2, 0)
app.store.load()
assert len(app.store.items) == 4, [i["asset_tag"] for i in app.store.items]
assert all(t in [i["asset_tag"] for i in app.store.items] for t in protetti), \
    "gli iPhone protetti sopravvivono anche alla reimportazione"
app.destroy()
print("RESET E EXPORT OK")
