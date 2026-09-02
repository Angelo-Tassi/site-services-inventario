"""I file di prova hanno la forma dei file veri.

Se il modello da compilare cambia colonne o ordine e i file di prova restano
indietro, il collaudo verifica una cosa che non esiste piu'. Questa suite lega
le due cose fra loro.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from inventario.excel_io import CAMPI_ESPORTAZIONE, TEMPLATE_FIELDS
from inventario.store import HEADERS, map_headers, rows_from_workbook

RADICE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CARTELLA = os.path.join(RADICE, "Collaudo")
# Il file di prova e' piu' ricco del modello: serve a provare che l'importazione
# riconosca tutte le colonne che puo' incontrare in un foglio vero.
ATTESE = ["Asset Tag", "Tipo", "Note", "Stato", "Modello/Descrizione",
          "Numero di serie"]
STANZE = [fixture.BAU, fixture.KIOSK, fixture.DR]

def foglio(nome):
    wb = load_workbook(os.path.join(CARTELLA, nome))
    try:
        ws = wb.worksheets[0]
        testa = [c for c in next(ws.iter_rows(values_only=True)) if c]
        larghezze = [ws.column_dimensions[get_column_letter(i)].width
                     for i in range(1, len(testa) + 1)]
        return testa, larghezze
    finally:
        wb.close()

# ---- il file regolare ha le colonne attese, nell'ordine dell'elenco
testa, larghezze = foglio("Inventario_di_prova.xlsx")
assert testa == ATTESE, testa

# ---- e ogni sua colonna e' riconosciuta dall'importazione: se un giorno un
# nome cambia senza che l'alias lo segua, il collaudo se ne accorge qui
riconosciute = map_headers(testa)
assert len(riconosciute) == len(testa), (riconosciute, testa)

# ---- contiene tutto quello che il modello propone, tranne la stanza, che nel
# file di prova la dicono le righe separatore
mancanti = (set(TEMPLATE_FIELDS) - {"stanza"}) - set(riconosciute.values())
assert not mancanti, mancanti
assert "stanza" in CAMPI_ESPORTAZIONE, "l'esportazione la stanza ce l'ha"

# ---- e nessuna colonna piu' stretta della propria intestazione
for nome, larghezza in zip(testa, larghezze):
    assert larghezza >= len(nome), (nome, larghezza)

# ---- il file con i difetti ha le stesse colonne, piu' quelle da ignorare
testa, _ = foglio("Inventario_di_prova_con_difetti.xlsx")
assert testa[:len(ATTESE)] == ATTESE, testa
assert testa[len(ATTESE):] == ["IMEI", "Costo", "Fornitore", "Centro di costo"], testa

# ---- e tutti e due si importano come dice il documento di collaudo
items, esito = rows_from_workbook(os.path.join(CARTELLA, "Inventario_di_prova.xlsx"),
                                  STANZE)
assert len(items) == 30, len(items)
assert esito["stanze_trovate"] == STANZE, esito["stanze_trovate"]
assert esito["scartate"] == 0 and esito["colonne_ignorate"] == [], esito

items, esito = rows_from_workbook(
    os.path.join(CARTELLA, "Inventario_di_prova_con_difetti.xlsx"), STANZE)
assert esito["colonne_ignorate"] == ["Costo", "Fornitore", "Centro di costo"], esito
assert esito["scartate"] == 1, esito          # la riga senza identificativo
assert esito["iphone"] == 1, esito            # il telefono ignorato
assert esito["senza_modello"] == 1, esito

print("COLLAUDO ALLINEATO OK")
