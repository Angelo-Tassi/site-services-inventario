"""Nei file prodotti non finiscono colonne che resterebbero vuote.

Un'esportazione non contiene iPhone, quindi non ha nessun IMEI da scrivere; una
stanza senza prestiti non ha prestiti. La stampa invece gli iPhone li include, e
li' l'IMEI serve. Le colonne portanti ci sono sempre, cosi' due file dello
stesso inventario si somigliano.
"""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from inventario import excel_io
from inventario.excel_io import (CAMPI_PORTANTI, TEMPLATE_FIELDS,
                                 campi_con_valore)
from inventario.store import ALL_FIELDS, InventoryStore, new_item

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
percorso = fixture.build()
store = InventoryStore(percorso, iphone_room=BAU)
store.add(new_item(tipo=fixture.TIPO_IPHONE, modello="Apple iPhone 14",
                   imei="356938035643809", restituito_da="M. Bianchi"))
items = store.load()
fuori = tempfile.mkdtemp()

def intestazioni(file_prodotto):
    wb = load_workbook(file_prodotto)
    try:
        for riga in wb.worksheets[0].iter_rows(values_only=True):
            if riga and riga[0] and str(riga[0]).strip() == "Asset Tag":
                return [c for c in riga if c]
        raise AssertionError("intestazioni non trovate in " + file_prodotto)
    finally:
        wb.close()

# ---- un file esportato dice che cosa abbiamo, dove sta e che cosa c'e' da
# sapere. Stesse quattro colonne ovunque.
ATTESE = ["Asset Tag", "Tipo", "Stanza", "Note"]
colonne = intestazioni(excel_io.export(items, os.path.join(fuori, "tutto.xlsx"),
                                       rooms=[BAU, KIOSK, DR]))
assert colonne == ATTESE, colonne

# ---- vale anche per la stanza dei prestiti: sono dati di consultazione,
# servono davanti all'elenco e non a chi riceve il file
kiosk = [i for i in items if i.get("stanza") == KIOSK]
assert any(i.get("prestato_a") for i in kiosk), "servono prestiti in corso"
colonne = intestazioni(excel_io.export(kiosk, os.path.join(fuori, "kiosk.xlsx"),
                                       rooms=[KIOSK], titolo=KIOSK))
assert colonne == ATTESE, colonne

# ---- e per il magazzino
dr = [i for i in items if i.get("stanza") == DR]
colonne = intestazioni(excel_io.export(dr, os.path.join(fuori, "dr.xlsx"),
                                       rooms=[DR], titolo=DR))
assert colonne == ATTESE, colonne

# ---- la stampa include gli iPhone, quindi l'IMEI le serve
colonne = intestazioni(excel_io.build_print_file(items, rooms=[BAU, KIOSK, DR]))
assert "IMEI" in colonne and "Restituito da" in colonne, colonne

# ---- il modello da compilare non ha mai i campi degli iPhone: non si importano
modello = os.path.join(fuori, "modello.xlsx")
excel_io.build_template(modello, [BAU, KIOSK, DR])
colonne = intestazioni(modello)
# il modello ha esattamente le colonne di un file esportato: si esporta, si
# corregge in Excel, si reimporta
assert colonne == ATTESE, colonne
assert excel_io.TEMPLATE_FIELDS == excel_io.CAMPI_ESPORTAZIONE
# le colonne che il modello non ha restano importabili: chi ha un foglio suo
# che le contiene lo carica lo stesso
from inventario.store import map_headers
riconosciute = map_headers(["Asset Tag", "Modello/Descrizione", "Numero di serie",
                            "Stato"])
assert sorted(riconosciute.values()) == ["asset_tag", "modello", "seriale", "stato"]

# ---- un file per stanza: tutti con la stessa forma
cartella = tempfile.mkdtemp()
scritti = excel_io.export_per_stanza(items, cartella, [BAU, KIOSK, DR])
assert len(scritti) == 3, scritti
for f in scritti:
    assert intestazioni(f) == ATTESE, (f, intestazioni(f))

# ---- da un'esportazione NON si ricostruisce un inventario: modello, seriale e
# note non ci sono. E' il motivo per cui la copia locale esiste.
from inventario.store import rows_from_workbook
riletti, _ = rows_from_workbook(scritti[0], [BAU, KIOSK, DR])
assert riletti, scritti[0]
assert all(not i["modello"] and not i["seriale"] for i in riletti)
assert all(i["asset_tag"] and i["stanza"] for i in riletti)
# le note invece viaggiano: sono quello che una riga ha di particolare
assert any(i["note"] for i in riletti), "le note devono sopravvivere all'export"

# ---- le colonne portanti restano anche quando non c'e' niente da scrivere
assert campi_con_valore([], list(ALL_FIELDS)) == CAMPI_PORTANTI
vuoti = [{"asset_tag": "IT-1", "tipo": "Laptop", "modello": "T14",
          "seriale": "", "stanza": BAU, "stato": "Disponibile", "note": ""}]
assert campi_con_valore(vuoti, list(ALL_FIELDS)) == CAMPI_PORTANTI

print("COLONNE ESPORTATE OK")
