"""Esportazione: scelta iniziale come per l'importazione, e fogli identificabili.

Il difetto segnalato: scegliendo un foglio per stanza usciva un foglio senza
titolo e senza colonna Stanza, impossibile da attribuire.
"""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from inventario import excel_io
from inventario.store import HEADERS, InventoryStore, new_item, rows_from_workbook
from inventario.ui import App, ExportOptionsDialog, nome_file

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
STANZE = [BAU, KIOSK, DR]
app = App(fixture.build()); app._initial_load()
app._run(lambda: app.store.add(new_item(tipo=app.iphone_type(), modello="iPhone 14",
                                        imei="356938035643809", restituito_da="M. B.")), "ok")
d = tempfile.mkdtemp()

# ------------------------------------------------ la finestra delle opzioni
dlg = ExportOptionsDialog(app, STANZE)
assert dlg.var_ambito.get() == "tutto" and dlg.var_forma.get() == "unico"
assert str(dlg.combo.cget("state")) == "disabled"
assert all(str(b.cget("state")) == "normal" for b in dlg.scelte_forma)
dlg.var_ambito.set("stanza"); dlg._aggiorna()
assert str(dlg.combo.cget("state")) == "readonly"
assert all("disabled" in str(b.cget("state")) for b in dlg.scelte_forma), \
    "con una stanza sola la forma non ha senso"
dlg.var_stanza.set(DR); dlg._ok()
assert dlg.result == {"stanza": DR, "forma": "unico"}, dlg.result

dlg = ExportOptionsDialog(app, STANZE)
dlg.var_forma.set("fogli"); dlg._ok()
assert dlg.result == {"stanza": None, "forma": "fogli"}

# ------------------------------------------------ un foglio per stanza: identificabile
percorso = os.path.join(d, "per_stanza.xlsx")
excel_io.export(app.store.items, percorso, group_by_room=True, rooms=STANZE)
wb = load_workbook(percorso)
assert sorted(wb.sheetnames) == sorted(STANZE), wb.sheetnames
for ws in wb.worksheets:
    assert ws["A1"].value == ws.title, (ws.title, ws["A1"].value)
    assert "Esportato il" in ws["A2"].value
    intestazioni = [c.value for c in ws[4]]
    assert HEADERS["stanza"] in intestazioni, "la colonna Stanza non deve sparire"
    colonna = intestazioni.index(HEADERS["stanza"])
    valori = {r[colonna] for r in ws.iter_rows(min_row=5, values_only=True) if r[0]}
    assert valori == {ws.title}, (ws.title, valori)
wb.close()
# e ogni foglio si reimporta da solo
righe, esito = rows_from_workbook(percorso, STANZE)
assert len(righe) == 13 and not any(i["imei"] for i in righe), len(righe)

# ------------------------------------------------ un file per stanza
cartella = tempfile.mkdtemp()
scritti = excel_io.export_per_stanza(app.store.items, cartella, STANZE)
assert len(scritti) == 3, scritti
attesi = {"Inventario_%s_" % nome_file(s) for s in STANZE}
for percorso in scritti:
    nome = os.path.basename(percorso)
    assert any(nome.startswith(a) for a in attesi), nome
    wb = load_workbook(percorso); ws = wb.active
    stanza = ws["A1"].value
    assert stanza in STANZE and ws.title == stanza, (ws.title, stanza)
    righe = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]
    assert len(righe) == 5 if stanza != DR else len(righe) == 3, (stanza, len(righe))
    wb.close()
# le stanze senza dispositivi non producono file
vuoto = InventoryStore(os.path.join(d, "vuoto.xlsx"), iphone_room=BAU)
vuoto.create_if_missing()
vuoto.add(new_item("IT-1", "Laptop", "T14", "SN1", KIOSK))
vuoto.load()
solo_uno = excel_io.export_per_stanza(vuoto.items, tempfile.mkdtemp(), STANZE)
assert len(solo_uno) == 1 and nome_file(KIOSK) in os.path.basename(solo_uno[0])

# ------------------------------------------------ una sola stanza
percorso = os.path.join(d, "solo_dr.xlsx")
items = [i for i in app.store.items if i["stanza"] == DR]
excel_io.export(items, percorso, rooms=[DR], titolo=DR)
wb = load_workbook(percorso); ws = wb.active
assert ws.title == DR and ws["A1"].value == DR
righe = [r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]
assert len(righe) == 3, len(righe)
wb.close()

# ------------------------------------------------ nessun iPhone, in nessuna forma
for percorso in [os.path.join(d, "per_stanza.xlsx"), os.path.join(d, "solo_dr.xlsx")] + scritti:
    wb = load_workbook(percorso)
    valori = [c.value for foglio in wb.worksheets for riga in foglio.iter_rows()
              for c in riga if c.value]
    assert "356938035643809" not in valori, percorso
    wb.close()
app.destroy()
print("OPZIONI EXPORT OK")
