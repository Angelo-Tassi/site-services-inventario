"""Nei file Excel prodotti nessuna colonna e' piu' stretta del suo contenuto.

Le larghezze erano fisse: il testo veniva tagliato o sbordava sulla cella
accanto, e chi apriva il file doveva allargare le colonne a mano una per una.
Vale per i file esportati, per la stampa e per il file dati, che si apre in
Excel come tutti gli altri.
"""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from inventario import excel_io
from inventario.store import (HEADERS, InventoryStore, LARGHEZZA_COLONNA_MAX,
                              LARGHEZZA_COLONNA_MIN, larghezza_colonna, new_item)

LUNGO = "Lenovo ThinkPad T14 Gen 5 - configurazione speciale reparto grafica"
NOTA = "Nota lunga scritta apposta per vedere se la colonna la segue davvero"

percorso = fixture.build()
store = InventoryStore(percorso, iphone_room=fixture.BAU)
store.add(new_item("IT-8000", "Laptop", LUNGO, "PF8000", fixture.BAU, NOTA))
items = store.load()
fuori = tempfile.mkdtemp()

def controlla(file_prodotto, etichetta):
    wb = load_workbook(file_prodotto)
    try:
        ws = wb.worksheets[0]
        testa = None
        for numero, riga in enumerate(ws.iter_rows(values_only=True), start=1):
            if riga and riga[0] and str(riga[0]).strip() == "Asset Tag":
                testa, prima = numero, riga
                break
        assert testa, etichetta
        for colonna, nome in enumerate([c for c in prima if c], start=1):
            larghezza = ws.column_dimensions[get_column_letter(colonna)].width
            assert larghezza, "%s: colonna %s senza larghezza" % (etichetta, nome)
            contenuto = max(len(str(ws.cell(row=r, column=colonna).value or ""))
                            for r in range(testa, ws.max_row + 1))
            assert larghezza >= contenuto, \
                "%s / %s: larga %.1f per un contenuto di %d" % (etichetta, nome,
                                                                larghezza, contenuto)
    finally:
        wb.close()

controlla(excel_io.export(items, os.path.join(fuori, "tutto.xlsx"),
                          rooms=[fixture.BAU, fixture.KIOSK, fixture.DR]), "esportazione")
controlla(excel_io.export([i for i in items if i.get("stanza") == fixture.KIOSK],
                          os.path.join(fuori, "kiosk.xlsx"), rooms=[fixture.KIOSK],
                          titolo=fixture.KIOSK), "esportazione di una stanza")
controlla(excel_io.build_print_file(items, rooms=[fixture.BAU]), "stampa")
controlla(percorso, "file dati")

# ---- una colonna vuota resta larga quanto il suo nome
assert larghezza_colonna(HEADERS["restituito_da"], []) >= len(HEADERS["restituito_da"])
assert larghezza_colonna("X", []) == LARGHEZZA_COLONNA_MIN

# ---- e nessuna diventa smisurata: un testo enorme non sfonda il foglio
assert larghezza_colonna("X", ["a" * 500]) == LARGHEZZA_COLONNA_MAX

# ---- il margine c'e': i caratteri non sono tutti larghi uguali
assert larghezza_colonna("Stanza", ["Magazzino Disaster Recovery"]) > \
    len("Magazzino Disaster Recovery")

print("LARGHEZZA FOGLI OK")
