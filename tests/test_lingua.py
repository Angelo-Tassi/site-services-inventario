"""Traduzione dell'interfaccia, dei file prodotti e ritorno all'italiano."""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from inventario import config, excel_io
from inventario import lingua as lang
from inventario.lingua import EN, INTESTAZIONI_EN, STATI_EN, T, intestazione
from inventario.store import (ALL_FIELDS, DISPONIBILE, HEADERS, InventoryStore,
                              STATI, new_item, rows_from_workbook)
from inventario.ui import App, ExportOptionsDialog, RoomsDialog, stato_canonico

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
STANZE = [BAU, KIOSK, DR]
d = tempfile.mkdtemp()

# ---- ogni intestazione e ogni stato hanno la loro traduzione
for campo in ALL_FIELDS:
    assert HEADERS[campo] in INTESTAZIONI_EN, campo
for stato in STATI + ["In prestito", "Da Rispedire", "Spedito al servizio telefonia"]:
    assert stato in STATI_EN, stato

# ---- l'italiano e' il testo sorgente: senza traduzione resta leggibile
lang.imposta("it")
assert T("Aggiungi") == "Aggiungi"
assert T("una frase mai tradotta") == "una frase mai tradotta"
lang.imposta("en")
assert T("Aggiungi") == "Add"
assert T("Reset inventario") == "Reset inventory"
assert T("una frase mai tradotta") == "una frase mai tradotta"
assert intestazione(HEADERS["seriale"]) == "Serial number"

# ---- l'interfaccia parte nella lingua salvata e si ricostruisce
lang.imposta("it")
percorso = fixture.build()
app = App(percorso); app._initial_load()
assert app.tree.heading("seriale")["text"].startswith("Numero di serie")
assert "Site Services" in app.title()
lang.imposta("en")
app.ricostruisci()
assert app.tree.heading("seriale")["text"].startswith("Serial number"), \
    app.tree.heading("seriale")["text"]
assert app.tree.heading("stato")["text"].startswith("Status")
assert app.view == "home" and len(app.visible) == 13
riga = app.tree.item("IT-0107", "values")
assert riga[app._columns().index("stato")] == "On loan", riga

# ---- lo stato si mostra tradotto ma si salva in italiano
assert stato_canonico("To be rebuilt", STATI) == "Da rebuildare"
assert stato_canonico("Da rebuildare", STATI) == "Da rebuildare"
app._run(lambda: app.store.set_stato("IT-0101", "Controllare"), "ok")
assert app._item_by_tag("IT-0101")["stato"] == "Controllare", "nel dato resta l'italiano"
assert app.tree.item("IT-0101", "values")[app._columns().index("stato")] == "To be checked"

# ---- la tendina della lingua nelle impostazioni
dlg = RoomsDialog(app, STANZE, ["Laptop"], [KIOSK], BAU, "en")
assert dlg.var_lingua.get() == "English"
dlg.var_lingua.set("Italiano"); dlg._ok()
assert dlg.result["lingua"] == "it", dlg.result

# ---- l'opzione di esportazione in inglese
lang.imposta("it")
app.ricostruisci()
dlg = ExportOptionsDialog(app, STANZE)
assert dlg.var_inglese.get() is False, "in italiano non e' spuntata di serie"
dlg.var_inglese.set(True); dlg._ok()
assert dlg.result["lingua"] == "en", dlg.result
lang.imposta("en")
dlg = ExportOptionsDialog(app, STANZE)
assert dlg.var_inglese.get() is True, "in inglese e' spuntata di serie"
dlg._cancel()
lang.imposta("it")

# ---- il file esportato in inglese, e il ritorno all'italiano
uscita = os.path.join(d, "export_en.xlsx")
excel_io.export(app.store.items, uscita, lingua="en")
wb = load_workbook(uscita); ws = wb.active
teste = [c.value for c in ws[1]]
from inventario.excel_io import CAMPI_ESPORTAZIONE
assert teste == [INTESTAZIONI_EN[HEADERS[f]] for f in CAMPI_ESPORTAZIONE], teste
colonna = teste.index("Status")
valori = {r[colonna] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
assert valori <= set(STATI_EN.values()), valori
assert "Available" in valori
wb.close()

righe, esito = rows_from_workbook(uscita, STANZE)
assert len(righe) == 13, len(righe)
assert {i["stato"] for i in righe} <= set(STATI + ["In prestito"]), \
    "rileggendo, gli stati tornano in italiano"
assert any(i["stato"] == DISPONIBILE for i in righe)
# l'esportazione porta quattro campi: il seriale non c'e' piu', e va bene cosi'
assert all(not i["seriale"] for i in righe), "l'export non porta il seriale"
assert all(i["asset_tag"] and i["stanza"] for i in righe)

# un file inglese si importa davvero
nuovo = InventoryStore(os.path.join(d, "Inventario.xlsx"), iphone_room=BAU)
nuovo.create_if_missing()
assert nuovo.import_items(righe, "merge")["aggiunti"] == 13
nuovo.load()
assert {i["stanza"] for i in nuovo.items} == set(STANZE)

# ---- esportazione per stanza in inglese
cartella = tempfile.mkdtemp()
scritti = excel_io.export_per_stanza(app.store.items, cartella, STANZE, lingua="en")
wb = load_workbook(scritti[0]); ws = wb.active
assert "Exported on" in ws["A2"].value, ws["A2"].value
assert ws["A1"].value in STANZE, "il nome della stanza non si traduce"
wb.close()

# ---- la lingua si ricorda, senza toccare la configurazione vera
finta_cartella = tempfile.mkdtemp()
config.app_dir = lambda: finta_cartella
config.load_language = fixture.load_language_reale   # qui la si prova sul serio
os.environ["APPDATA"] = tempfile.mkdtemp()
assert config.load_language() == "it", "senza preferenza salvata si parte in italiano"
config.save_language("en")
assert config.load_language() == "en"
config.save_language("it")
assert config.load_language() == "it"
app.destroy()
print("LINGUA OK")
