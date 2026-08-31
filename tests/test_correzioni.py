"""I quattro punti segnalati: tipo prima di tutto, ritorno alla home,
ordine cronologico, iPhone senza seriale ne' prestiti."""
import os, sys, time
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from inventario.ui import (ACTION_COLUMN, AddChoiceDialog, App, COLONNE_NON_IPHONE,
                           ItemDialog, TypeChoiceDialog, chiave_ordinamento)
from inventario.store import InventoryStore, new_item, valore_visibile
from inventario import excel_io
from openpyxl import load_workbook

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
app = App(fixture.build()); app._initial_load()
TIPO = app.iphone_type()

# ---------------------------------------------- 1. prima si sceglie il tipo
scelta = TypeChoiceDialog(app, app.cfg["types"], app.tipo_predefinito())
assert list(scelta.winfo_children()[0].winfo_children()[2].cget("values")) == app.cfg["types"]
assert scelta.var_tipo.get() == "Laptop", scelta.var_tipo.get()
scelta.var_tipo.set(TIPO); scelta._ok()
assert scelta.result == TIPO
# nel contenitore iPhone la tendina propone gia' l'iPhone
app.show_iphones()
assert app.tipo_predefinito() == TIPO
app.show_home()
assert app.tipo_predefinito() == "Laptop"
# solo dopo si sceglie come inserirlo, con i campi del tipo scelto
assert AddChoiceDialog(app, iphone=True).title() == "Aggiungi iPhone"
assert AddChoiceDialog(app, iphone=False).title() == "Aggiungi dispositivo"

# ---------------------------------------------- 2. Home torna davvero alla home
app.show_iphones()
assert app.view == "type" and app.var_type.get() == TIPO
app.show_home()
assert app.view == "home", app.view
assert app.var_type.get() == "Tutti", app.var_type.get()
assert app.var_room.get() == "Tutte"
assert len(app.visible) == len(app.store.items), (len(app.visible), len(app.store.items))
# e lo stesso vale passando da una stanza
app.show_room(KIOSK); app.show_home()
assert len(app.visible) == len(app.store.items)

# ---------------------------------------------- 3. il piu' recente in cima
assert app.sort_field == "modificato_il" and app.sort_reverse is True
time.sleep(1.1)
app._run(lambda: app.store.add(new_item("IT-0900", "Laptop", "T14 Gen 5", "PFNEW1", DR)), "ok")
assert app.visible[0]["asset_tag"] == "IT-0900", [i["asset_tag"] for i in app.visible[:3]]
time.sleep(1.1)   # secondi diversi: l ordine deve essere certo
app._run(lambda: app.store.add(new_item("IT-0901", "Tablet",
                                        "Dell Latitude 7320 Detachable", "8HNEW", BAU)), "ok")
assert app.visible[0]["asset_tag"] == "IT-0901", [i["asset_tag"] for i in app.visible[:3]]
# anche dentro una stanza
app.show_room(BAU)
assert app.visible[0]["asset_tag"] == "IT-0901"
app.show_home()

# le date si ordinano per data, non per testo
a = {"modificato_il": "09/12/2026 08:00"}
b = {"modificato_il": "10/01/2027 08:00"}
assert chiave_ordinamento(a, "modificato_il") < chiave_ordinamento(b, "modificato_il")
assert chiave_ordinamento({"modificato_il": ""}, "modificato_il") \
    < chiave_ordinamento(a, "modificato_il")
# il testo resta alfabetico e parte dalla A
app.sort_by("modello")
assert app.sort_reverse is False
app.sort_by("spedito_il")
assert app.sort_reverse is True, "le date partono dalla piu' recente"

# ---------------------------------------------- 4. iPhone senza seriale ne' prestiti
app._run(lambda: app.store.add(new_item(tipo=TIPO, modello="Apple iPhone 14",
                                        imei="356938035643809", restituito_da="M. B.",
                                        seriale="RESIDUO", prestato_a="Tizio",
                                        prestato_il="01/01/2026 10:00")), "ok")
tel = app._item_by_tag("356938035643809")
assert tel["seriale"] == "" and tel["prestato_a"] == "" and tel["prestato_il"] == ""
# e nel loro contenitore quelle colonne non compaiono proprio
app.show_iphones()
colonne = app._columns()
for campo in COLONNE_NON_IPHONE:
    assert campo not in colonne, (campo, colonne)
assert "imei" in colonne and "restituito_da" in colonne
assert ACTION_COLUMN in colonne
# nelle altre schermate restano, servono a laptop e tablet
app.show_room(KIOSK)
for campo in COLONNE_NON_IPHONE:
    assert campo in app._columns(), campo

# ---------------------------------------------- 5. nessun asset tag sugli iPhone
assert "asset_tag" in COLONNE_NON_IPHONE
assert valore_visibile(tel, "asset_tag") == "", "non si mostra mai"
assert valore_visibile(tel, "imei") == "356938035643809"
assert valore_visibile(app._item_by_tag("IT-0900"), "asset_tag") == "IT-0900"

# nell'elenco generale la colonna dell'asset tag resta vuota per un telefono,
# e l'IMEI non c'e' proprio: si guarda nel contenitore Iphone
app.show_home()
colonne = app._columns()
riga = app.tree.item("356938035643809", "values")
assert riga[colonne.index("asset_tag")] == "", riga
assert "imei" not in colonne, colonne
app.show_iphones()
colonne = app._columns()
riga = app.tree.item("356938035643809", "values")
assert riga[colonne.index("imei")] == "356938035643809"
assert "asset_tag" not in colonne, colonne
app.show_home()

# la scheda di un iPhone non ha il campo Asset Tag
scheda = ItemDialog(app, app.cfg["rooms"], app.cfg["types"], tel,
                    iphone_room=app.iphone_room(), stati=app.cfg["states"])
assert "Asset Tag" not in [l for l, _v, _w in scheda.required]
scheda._cancel()

# non compare nemmeno nel file dati, ma l'identita' sopravvive alla rilettura
wb = load_workbook(app.store.path); ws = wb.active
intestazioni = [c.value for c in ws[1]]
colonna_tag = intestazioni.index("Asset Tag")
colonna_imei = intestazioni.index("IMEI")
telefoni = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if r[colonna_imei] == "356938035643809"]
assert len(telefoni) == 1 and telefoni[0][colonna_tag] in (None, ""), telefoni
wb.close()
app.store.load()
assert app._item_by_tag("356938035643809") is not None, "l IMEI ricostruisce la chiave"
assert app.store.set_note("356938035643809", "riletto") is True

# nemmeno nella stampa, dove gli iPhone ci sono
stampa = excel_io.build_print_file(app.store.items, rooms=app.cfg["rooms"])
wb = load_workbook(stampa)
ws = wb.active
riga_intestazioni = next(r for r in ws.iter_rows(values_only=True) if r and r[0] == "Asset Tag")
indice = list(riga_intestazioni).index("Asset Tag")
valori = [r[indice] for r in ws.iter_rows(min_row=2, values_only=True) if r]
assert "356938035643809" not in [v for v in valori if v], valori
wb.close()
app.destroy()
print("CORREZIONI OK")
