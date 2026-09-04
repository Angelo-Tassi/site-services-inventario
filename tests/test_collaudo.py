"""Ripete sui file di Collaudo/ gli scenari descritti nelle sue istruzioni.

Se il comportamento del programma cambia, questa suite fallisce e le istruzioni
di collaudo vanno riscritte: non devono mai promettere cose che non succedono.
"""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from inventario import excel_io
from inventario.store import InventoryStore, new_item, rows_from_workbook

RADICE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COLLAUDO = os.path.join(RADICE, "Collaudo")
COMPLETO = os.path.join(COLLAUDO, "Inventario_di_prova.xlsx")
DIFETTI = os.path.join(COLLAUDO, "Inventario_di_prova_con_difetti.xlsx")
BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
STANZE = [BAU, KIOSK, DR]

assert os.path.exists(COMPLETO) and os.path.exists(DIFETTI), \
    "rigenera i file con tests/genera_file_di_prova.py"
istruzioni = open(os.path.join(COLLAUDO, "README.md"), encoding="utf-8").read()
assert "Come testare l'importazione" in istruzioni

def inventario_vuoto():
    p = os.path.join(tempfile.mkdtemp(), "Inventario.xlsx")
    s = InventoryStore(p, iphone_room=BAU)
    s.create_if_missing()
    return s

def per_stanza(store):
    conteggi = {}
    for i in store.items:
        conteggi[i["stanza"]] = conteggi.get(i["stanza"], 0) + 1
    return conteggi

# ---------------------------------------------------- 1. caricamento iniziale
items, esito = rows_from_workbook(COMPLETO, STANZE)
assert len(items) == 30, len(items)
assert esito["da_tag"] == 30, esito
assert esito["scartate"] == 0 and esito["colonne_ignorate"] == [], esito
s = inventario_vuoto()
r = s.import_items(items, "merge")
assert (r["aggiunti"], r["gia_presenti"]) == (30, []), r
s.load()
assert per_stanza(s) == {BAU: 10, KIOSK: 10, DR: 10}, per_stanza(s)
# i tablet Dell ci sono e si distinguono dai laptop
tablet = [i for i in s.items if i["tipo"] == "Tablet"]
assert len(tablet) == 9 and all("Dell" in i["modello"] for i in tablet), len(tablet)

# ---------------------------------------------------- 2. l'unione non duplica
# Sono tutti gia' in inventario: nessuno entra, nessuno viene riscritto, e per
# ognuno si dice in che stanza sta quello che c'e' gia'.
r = s.import_items(items, "merge")
assert r["aggiunti"] == 0, r
assert len(r["gia_presenti"]) == 30, len(r["gia_presenti"])
assert {v["stanza"] for v in r["gia_presenti"]} == {BAU, KIOSK, DR}, r["gia_presenti"]
s.load()
assert per_stanza(s) == {BAU: 10, KIOSK: 10, DR: 10}

# ---------------------------------------------------- 3. sostituire una stanza
# Qui si passano tutte e 30 le righe forzando la stanza Digital Kiosk: dentro il
# programma non succede mai - "Una sola stanza" filtra prima le righe di quella
# stanza - ma serve a controllare che cosa fa l'archivio se ci provi.
# Le 10 del Kiosk vengono cancellate e rimesse; le 20 delle altre due stanze
# sono gia' in inventario e vengono saltate: un'importazione dentro una stanza
# non trascina li' dispositivi registrati altrove.
r = s.import_items(items, "replace", KIOSK)
assert r["eliminati"] == 10, r
assert r["aggiunti"] == 10, r
assert len(r["gia_presenti"]) == 20, len(r["gia_presenti"])
s.load()
assert per_stanza(s) == {BAU: 10, KIOSK: 10, DR: 10}, per_stanza(s)
assert r["copia"] and os.path.exists(r["copia"])
assert os.path.basename(os.path.dirname(r["copia"])) == "Backup", r["copia"]

# ---------------------------------------------------- 4. sostituire tutto
r = s.import_items(items, "replace", None)
s.load()
assert r["eliminati"] == 30, r
assert per_stanza(s) == {BAU: 10, KIOSK: 10, DR: 10}, per_stanza(s)

# ---------------------------------------------------- 5. il file con i difetti
items2, esito2 = rows_from_workbook(DIFETTI, STANZE)
assert len(items2) == 3, items2
assert esito2["scartate"] == 1, esito2
assert esito2["iphone"] == 1, esito2
assert esito2["senza_modello"] == 1, esito2
assert esito2["colonne_ignorate"] == ["Costo", "Fornitore", "Centro di costo"], esito2
assert esito2["da_tag"] == 3, esito2
prima = len(s.items)
s.import_items(items2, "merge")
s.load()
assert len(s.items) == prima + 3
nuovi = {i["asset_tag"]: i for i in s.items if i["asset_tag"].startswith("IT-BAU-9")
         or i["asset_tag"].startswith("IT-KSK-9")}
assert set(nuovi) == {"IT-BAU-901", "IT-BAU-902", "IT-KSK-903"}, set(nuovi)
assert nuovi["IT-BAU-901"]["stanza"] == BAU and nuovi["IT-KSK-903"]["stanza"] == KIOSK
assert nuovi["IT-BAU-902"]["modello"] == "", "la riga senza modello entra vuota"
assert not any(i["imei"] for i in s.items), "l'iPhone del file non deve entrare"

# ---------------------------------------------------- 6. reset e ricarica
s.add(new_item(tipo="Iphone", modello="Apple iPhone 14", imei="351111111111111",
               restituito_da="Collaudo"))
eliminati, tenuti, copia = s.reset()
s.load()
assert tenuti == 1 and [i["asset_tag"] for i in s.items] == ["351111111111111"]
s.import_items(items, "merge")
s.load()
assert len(s.items) == 31, len(s.items)
assert any(i["asset_tag"] == "351111111111111" for i in s.items), \
    "l'iPhone inserito a mano sopravvive al reset e alla ricarica"

# ---------------------------------------------------- 7. esportazione di stanza
uscita = os.path.join(tempfile.mkdtemp(), "Inventario_Digital_Kiosk_20260830.xlsx")
excel_io.export([i for i in s.items if i["stanza"] == KIOSK], uscita,
                rooms=[KIOSK], titolo=KIOSK)
wb = load_workbook(uscita); ws = wb.active
assert ws.title == KIOSK and ws["A1"].value == KIOSK
assert "Esportato il" in ws["A2"].value
valori = [r for r in ws.iter_rows(min_row=5, values_only=True) if r and r[0]]
assert len(valori) == 10, len(valori)
wb.close()
usciti, _ = rows_from_workbook(uscita, STANZE)
assert all(i["stanza"] == KIOSK for i in usciti)
assert not any(i["imei"] for i in usciti), "nessun iPhone nelle esportazioni"
print("COLLAUDO OK")
