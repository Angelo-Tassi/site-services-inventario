"""Il modello vuoto di importazione, e l'esclusione degli iPhone da import/export."""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fixture
from openpyxl import load_workbook
from inventario import excel_io
from inventario.store import (HEADERS, InventoryStore, new_item, rows_from_workbook)

BAU, KIOSK, DR = fixture.BAU, fixture.KIOSK, fixture.DR
STANZE = [BAU, KIOSK, DR]
d = tempfile.mkdtemp()
modello = os.path.join(d, "Modello_inventario.xlsx")
excel_io.build_template(modello, STANZE)

wb = load_workbook(modello)
assert wb.sheetnames == ["Inventario", "Istruzioni"], wb.sheetnames
ws = wb["Inventario"]

# ---- solo le colonne che servono a laptop e tablet
assert [c.value for c in ws[1]] == [HEADERS[f] for f in excel_io.TEMPLATE_FIELDS]
# le tendine devono stare sulle colonne di tipo e stato, comunque siano ordinate
from openpyxl.utils import get_column_letter
attese = {get_column_letter(excel_io.TEMPLATE_FIELDS.index(c) + 1) for c in ("tipo", "stato")}
trovate = {str(dv.sqref).split("2:")[0] for dv in ws.data_validations.dataValidation}
assert trovate == attese, (trovate, attese)
# e nessuna colonna piu' stretta della propria intestazione
for i, campo in enumerate(excel_io.TEMPLATE_FIELDS, start=1):
    larghezza = ws.column_dimensions[get_column_letter(i)].width
    assert larghezza >= len(HEADERS[campo]), (campo, larghezza)
assert "IMEI" not in [c.value for c in ws[1]]
assert "Stanza" not in [c.value for c in ws[1]], "la stanza arriva dai separatori"

# ---- un separatore per stanza, con righe libere sotto
separatori = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)
              if ws.cell(row=r, column=1).value]
assert separatori == [s.upper() for s in STANZE], separatori
assert ws.max_row == 1 + len(STANZE) * (1 + excel_io.RIGHE_PER_STANZA)

# ---- tendine su Tipo e Stato
valori = [v.formula1 for v in ws.data_validations.dataValidation]
assert any("Laptop" in v and "Tablet" in v and "Iphone" not in v for v in valori), valori
assert any("Disponibile" in v for v in valori), valori
wb.close()

# ---- il modello si importa senza errori, e a vuoto non porta niente
items, esito = rows_from_workbook(modello, STANZE)
assert items == [] and esito["scartate"] == 0, (items, esito)

# ---- compilato, finisce nelle stanze giuste
wb = load_workbook(modello)
ws = wb["Inventario"]
ws.cell(row=3, column=1, value="IT-0101"); ws.cell(row=3, column=2, value="Laptop")
ws.cell(row=3, column=3, value="Lenovo ThinkPad T14 Gen 4")
ws.cell(row=3, column=4, value="PF4A1B2C"); ws.cell(row=3, column=5, value="Disponibile")
riga_kiosk = 2 + (1 + excel_io.RIGHE_PER_STANZA) + 1
ws.cell(row=riga_kiosk, column=1, value="IT-0106"); ws.cell(row=riga_kiosk, column=2, value="Tablet")
ws.cell(row=riga_kiosk, column=3, value="Dell Latitude 7320 Detachable")
ws.cell(row=riga_kiosk, column=4, value="8H2KLM3")
compilato = os.path.join(d, "compilato.xlsx")
wb.save(compilato); wb.close()

items, esito = rows_from_workbook(compilato, STANZE)
assert len(items) == 2, items
assert {i["asset_tag"]: i["stanza"] for i in items} == {"IT-0101": BAU, "IT-0106": KIOSK}
assert esito["da_tag"] == 2 and esito["iphone"] == 0

p = os.path.join(d, "Inventario.xlsx")
s = InventoryStore(p, iphone_room=BAU); s.create_if_missing()
assert s.import_items(items, "merge")["aggiunti"] == 2
s.load(); assert len(s.items) == 2

# ---- gli iPhone restano fuori da import ed export
s.add(new_item(tipo="Iphone", modello="Apple iPhone 14", imei="356938035643809",
               restituito_da="M. B."))
s.load(); assert len(s.items) == 3
esportato = os.path.join(d, "export.xlsx")
excel_io.export(s.items, esportato)
usciti, _ = rows_from_workbook(esportato, STANZE)
assert len(usciti) == 2, [i["asset_tag"] for i in usciti]
assert not any(i["tipo"].lower() == "iphone" for i in usciti)

# per stanza: nemmeno un foglio con il telefono
per_stanza = os.path.join(d, "per_stanza.xlsx")
excel_io.export(s.items, per_stanza, group_by_room=True, rooms=STANZE)
wb = load_workbook(per_stanza)
valori = [c.value for foglio in wb.worksheets for riga in foglio.iter_rows()
          for c in riga if c.value]
assert "356938035643809" not in valori
wb.close()

# la sostituzione rifa' l'inventario dal file ma non cancella gli iPhone
assert s.import_items(items, "replace")["aggiunti"] == 2
s.load()
assert len(s.items) == 3, [i["asset_tag"] for i in s.items]
assert any(i["asset_tag"] == "356938035643809" for i in s.items)
print("MODELLO OK")
