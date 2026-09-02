#!/usr/bin/env python3
"""Rigenera i file di prova in Collaudo/.

Uso:  .venv/bin/python tests/genera_file_di_prova.py
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from inventario.excel_io import TEMPLATE_FIELDS
from inventario.store import HEADERS, larghezza_colonna

RADICE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CARTELLA = os.path.join(RADICE, "Collaudo")

# I file di prova sono piu' ricchi del modello da compilare, e devono esserlo:
# il modello contiene il minimo per caricare dei dispositivi, questi servono a
# provare che l'importazione riconosca tutte le colonne che puo' incontrare in
# un foglio vero. L'ordine e' quello dell'elenco nel programma.
CAMPI = ["asset_tag", "tipo", "note", "stato", "modello", "seriale"]
assert set(TEMPLATE_FIELDS) - {"stanza"} <= set(CAMPI), \
    "il file di prova deve contenere almeno le colonne del modello"
INTESTAZIONI = [HEADERS[c] for c in CAMPI]

LAPTOP = ["Lenovo ThinkPad T14 Gen 4", "Lenovo ThinkPad T14 Gen 5"]
TABLET = ["Dell Latitude 7320 Detachable", "Dell Latitude 7230 Rugged Extreme"]

STANZE = [
    ("SITE SERVICES BAU", "BAU", 100),
    ("DIGITAL KIOSK", "KSK", 200),
    ("MAGAZZINO DISASTER RECOVERY", "DRC", 300),
]

STATI = ["Disponibile", "Disponibile", "Disponibile", "In attesa ritiro",
         "Guasto in attesa tecnico", "Da rebuildare", "Controllare",
         "Disponibile", "Disponibile", "Disponibile"]

NOTE = ["", "Postazione 1", "", "Batteria da sostituire", "", "Con custodia",
        "", "In verifica", "", "Scorta"]


def _dispositivo(prefisso, base, indice):
    """Sette laptop e tre tablet per stanza, con seriali plausibili."""
    numero = base + indice
    if indice <= 7:
        modello = LAPTOP[indice % len(LAPTOP)]
        seriale = "PF%d%s%02d" % (4 + indice % 2, prefisso[:2], numero % 100)
        tipo = "Laptop"
    else:
        modello = TABLET[indice % len(TABLET)]
        seriale = "%dH%s%03d" % (4 + indice % 5, prefisso[:1], numero % 1000)
        tipo = "Tablet"
    valori = {"asset_tag": "IT-%s-%03d" % (prefisso, numero), "tipo": tipo,
              "modello": modello, "seriale": seriale,
              "stato": STATI[indice - 1], "note": NOTE[indice - 1]}
    return [valori[c] for c in CAMPI]


def _intesta(ws):
    ws.append(INTESTAZIONI)
    for cella in ws[1]:
        cella.font = Font(bold=True, color="FFFFFF")
        cella.fill = PatternFill("solid", fgColor="1F4E79")
        cella.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    _larghezze(ws)


def _larghezze(ws):
    """Colonne larghe quanto serve, come in tutti gli altri file prodotti."""
    from openpyxl.utils import get_column_letter
    for colonna in range(1, ws.max_column + 1):
        valori = [ws.cell(row=r, column=colonna).value
                  for r in range(2, ws.max_row + 1)]
        titolo = str(ws.cell(row=1, column=colonna).value or "")
        ws.column_dimensions[get_column_letter(colonna)].width = larghezza_colonna(
            titolo, valori)


def _separatore(ws, testo):
    riga = ws.max_row + 1
    cella = ws.cell(row=riga, column=1, value=testo)
    cella.font = Font(bold=True, color="1F4E79")
    cella.fill = PatternFill("solid", fgColor="DCE6F1")
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga,
                   end_column=len(INTESTAZIONI))


def inventario_completo(percorso):
    """Trenta dispositivi, dieci per stanza, divisi dai separatori."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    _intesta(ws)
    for tag, prefisso, base in STANZE:
        _separatore(ws, tag)
        for indice in range(1, 11):
            ws.append(_dispositivo(prefisso, base, indice))
    _larghezze(ws)
    wb.save(percorso)
    wb.close()
    return percorso


def inventario_con_difetti(percorso):
    """Lo stesso file, con dentro i casi che il programma deve segnalare."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    extra = ["IMEI", "Costo", "Fornitore", "Centro di costo"]
    ws.append(INTESTAZIONI + extra)
    for cella in ws[1]:
        cella.font = Font(bold=True, color="FFFFFF")
        cella.fill = PatternFill("solid", fgColor="1F4E79")

    def riga(**valori):
        base = dict.fromkeys(CAMPI, "")
        base.update((k, v) for k, v in valori.items() if k in base)
        return ([base[c] for c in CAMPI]
                + [valori.get("imei", ""), valori.get("costo", ""),
                   valori.get("fornitore", ""), valori.get("cc", "")])

    _separatore(ws, "BAU")
    ws.append(riga(asset_tag="IT-BAU-901", tipo="Laptop",
                   modello="Lenovo ThinkPad T14 Gen 5", seriale="PF4BAU01",
                   stato="Disponibile", note="riga regolare",
                   costo=1200, fornitore="Dell Italia", cc="CC-01"))
    ws.append(riga(asset_tag="IT-BAU-902", tipo="Laptop", seriale="PF4BAU02",
                   stato="Disponibile", note="senza modello",
                   costo=1150, fornitore="Dell Italia", cc="CC-01"))
    ws.append([None] * (len(CAMPI) + len(extra)))
    _separatore(ws, "KIOSK")
    ws.append(riga(asset_tag="IT-KSK-903", tipo="Tablet",
                   modello="Dell Latitude 7320 Detachable", seriale="8HK903",
                   stato="Controllare", costo=900, fornitore="Dell Italia", cc="CC-02"))
    ws.append(riga(tipo="Laptop", modello="senza identificativo",
                   note="verra' scartata", costo=800, fornitore="Dell Italia",
                   cc="CC-02"))
    ws.append(riga(tipo="Iphone", modello="Apple iPhone 14", note="verra' ignorato",
                   imei="356938035643809", costo=1000, fornitore="Apple", cc="CC-03"))
    _larghezze(ws)
    wb.save(percorso)
    wb.close()
    return percorso


def main():
    if not os.path.isdir(CARTELLA):
        os.makedirs(CARTELLA)
    uno = inventario_completo(os.path.join(CARTELLA, "Inventario_di_prova.xlsx"))
    due = inventario_con_difetti(os.path.join(CARTELLA, "Inventario_di_prova_con_difetti.xlsx"))
    for percorso in (uno, due):
        print("scritto:", os.path.relpath(percorso, RADICE))


if __name__ == "__main__":
    main()
