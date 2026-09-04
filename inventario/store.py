"""Archivio dati dell'inventario.

I dati risiedono in un unico file .xlsx sulla cartella di rete: e' allo stesso
tempo il database e l'inventario apribile in Excel.

Sicurezza sugli accessi concorrenti:
  * ogni modifica avviene dentro un lock esclusivo (file .lock creato con
    O_CREAT|O_EXCL, che su SMB e' atomico);
  * dentro il lock i dati vengono sempre riletti da disco e l'operazione
    (aggiunta / modifica / eliminazione) viene riapplicata sui dati freschi,
    quindi due utenti che lavorano su schede diverse non si sovrascrivono;
  * la scrittura avviene su un file temporaneo nella stessa cartella e poi
    os.replace(), quindi il file non resta mai a meta'.
"""

import calendar
import getpass
import json
import os
import platform
import shutil
import socket
import tempfile
import time
import uuid
import zipfile
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook

from .lingua import T

SHEET_NAME = "Inventario"

FIELDS = ["asset_tag", "tipo", "modello", "seriale", "imei", "restituito_da",
          "stanza", "stato", "prestato_a", "prestato_il", "spedito_il", "note"]

# Per un iPhone l'identita' e' l'IMEI, non l'asset tag aziendale.
TIPO_IPHONE = "iphone"

DISPONIBILE = "Disponibile"
# Uno stato che dice quello che e' successo, non quello che manca: chi legge
# l'elenco vuole sapere che il dispositivo e' in prestito, non che "non e'
# disponibile". Viene ricalcolato dal prestito a ogni lettura, quindi gli
# inventari scritti prima si allineano da soli.
NON_DISPONIBILE = "In prestito"
# Gli iPhone in nostro possesso sono sempre in attesa di essere rispediti.
DA_RISPEDIRE = "Da Rispedire"
# ...finche' non partono davvero per il servizio telefonia.
SPEDITO = "Spedito al servizio telefonia"

# Un dispositivo spedito resta consultabile in inventario per tre mesi.
MESI_CONSERVAZIONE = 3

TESTO_SPEDIZIONE = (
    "Il dispositivo e' stato rispedito al servizio telefonia il %s. "
    "Resta in inventario per consultazione fino al %s, data dalla quale potra' "
    "essere eliminato."
)

# Stati scegliibili a mano. Gli altri due sono automatici: NON_DISPONIBILE
# mentre c'e' un prestito in corso, DA_RISPEDIRE per gli iPhone.
STATI = [
    DISPONIBILE,
    "In attesa ritiro",
    "Guasto in attesa tecnico",
    "Da rebuildare",
    "Controllare",
]
AUDIT_FIELDS = ["modificato_il", "modificato_da"]
ALL_FIELDS = FIELDS + AUDIT_FIELDS

HEADERS = {
    "asset_tag": "Asset Tag",
    "tipo": "Tipo",
    "modello": "Modello/Descrizione",
    "seriale": "Numero di serie",
    "imei": "IMEI",
    "restituito_da": "Restituito da",
    "stanza": "Stanza",
    "stato": "Stato",
    "prestato_a": "In prestito a",
    "prestato_il": "Prestato il",
    "spedito_il": "Spedito il",
    "note": "Note",
    "modificato_il": "Ultima modifica",
    "modificato_da": "Modificato da",
}

# Intestazioni accettate in importazione (minuscole, senza spazi doppi).
HEADER_ALIASES = {
    "asset_tag": ["asset tag", "assettag", "asset", "tag", "etichetta", "inventario"],
    "tipo": ["tipo", "tipologia", "categoria", "device type", "type"],
    "modello": ["modello", "model", "modello/descrizione", "model/description",
                "descrizione", "description", "dispositivo", "device"],
    "seriale": ["numero di serie", "seriale", "serial", "serial number", "s/n",
                "sn", "matricola", "service tag"],
    "imei": ["imei", "imei/meid", "meid", "codice imei"],
    "restituito_da": ["restituito da", "proprietario", "consegnato da",
                      "riconsegnato da", "owner", "returned by"],
    "stanza": ["stanza", "room", "locale", "ubicazione", "posizione", "location"],
    "stato": ["stato", "status", "disponibilita", "disponibilita'"],
    "prestato_a": ["in prestito a", "prestato a", "prestito", "assegnato a",
                   "utilizzatore", "borrower", "assigned to", "on loan to"],
    "prestato_il": ["prestato il", "data prestito", "in prestito dal",
                    "loan date", "borrowed on", "lent on"],
    "spedito_il": ["spedito il", "data spedizione", "rispedito il",
                   "shipped on"],
    "note": ["note", "nota", "commenti", "notes"],
    "modificato_il": ["ultima modifica", "modificato il", "data", "last change"],
    "modificato_da": ["modificato da", "utente", "changed by"],
}

# Parole ignorate quando si ricavano i tag dai nomi delle stanze.
_PAROLE_VUOTE = {"DEL", "DELLA", "DEI", "DELLE", "DI", "DA", "IN", "PER", "LA", "IL"}


def tag_stanze(rooms):
    """Tag riconosciuti in importazione, ricavati dai nomi delle stanze.

    Per ogni stanza vale il nome completo, piu' ogni singola parola che sia
    inequivocabile: "Digital Kiosk" si scrive anche solo KIOSK, "Site Services
    BAU" anche solo BAU, "Magazzino Disaster Recovery" anche solo DISASTER.
    """
    tags = {}
    parole = {}
    for stanza in rooms:
        nome = clean(stanza).upper()
        if not nome:
            continue
        tags[nome] = stanza
        for parola in nome.split():
            if len(parola) < 3 or parola in _PAROLE_VUOTE:
                continue
            parole.setdefault(parola, set()).add(stanza)
    for parola, stanze in parole.items():
        if len(stanze) == 1 and parola not in tags:
            tags[parola] = list(stanze)[0]
    return tags


def riga_tag(row, tags):
    """Se la riga e' un separatore di stanza ritorna la stanza, altrimenti None.

    Un separatore e' una riga con una sola cella scritta, che contiene il tag.
    """
    valori = [clean(c) for c in (row or ()) if clean(c)]
    if len(valori) != 1:
        return None
    testo = valori[0].upper().strip(" :-\u2013\u2014.\t")
    return tags.get(testo)


def separatore_con_avanzi(row, mapping, tags):
    """Riconosce un separatore anche quando la riga porta celle di troppo.

    La forma pulita - una sola cella scritta - la vede gia' riga_tag. Qui si
    prendono i fogli meno ordinati: vale come separatore la riga in cui la
    casella dell'identificativo contiene il nome di una stanza e nessun altro
    campo del dispositivo e' pieno. Un dispositivo vero non si chiama come una
    stanza, quindi il rischio di scambiarli e' nullo; il danno di non
    riconoscere un separatore, invece, e' che tutto l'inventario finisce senza
    stanza e il nome della stanza diventa un dispositivo.
    """
    campi = {}
    for idx, field in mapping.items():
        campi.setdefault(field, idx)
    idx = campi.get("asset_tag", campi.get("imei"))
    if idx is None or idx >= len(row):
        return None
    testo = clean(row[idx]).upper().strip(" :-\u2013\u2014.\t")
    if testo not in tags:
        return None
    for field in ("tipo", "modello", "seriale", "imei", "stato"):
        altro = campi.get(field)
        if altro is not None and altro != idx and altro < len(row) \
                and clean(row[altro]):
            return None
    return tags[testo]


def righe_separatore(items, rooms):
    """I dispositivi che in realta' sono righe separatore di stanza.

    Se ne compaiono, il file aperto non e' un inventario ma un foglio da
    importare: l'inventario vero non contiene una riga il cui identificativo e'
    il nome di una stanza. E' l'unico modo, guardando i dati, di accorgersi che
    e' stato scambiato un file per l'altro.
    """
    tags = tag_stanze(rooms or [])
    trovate = []
    for it in items or []:
        testo = clean(it.get("asset_tag")).upper().strip(" :-\u2013\u2014.\t")
        if testo in tags and not clean(it.get("tipo")) and not clean(it.get("modello")):
            trovate.append(it.get("asset_tag"))
    return trovate


def sembra_un_foglio_da_importare(path, rooms):
    """Guarda un file gia' esistente e dice se e' un foglio da importare.

    Ritorna (si_o_no, motivo). Serve prima di adottare un file come inventario:
    scegliere per sbaglio un file di prova significa ritrovarsi le righe
    separatore in elenco come se fossero dispositivi, e nessuna stanza.
    """
    try:
        store = InventoryStore(path)
        items = store.load()
    except Exception:
        return False, ""
    separatori = righe_separatore(items, rooms)
    if separatori:
        return True, ("Contiene righe separatore di stanza (%s): sono le righe "
                      "che dividono un foglio da importare."
                      % ", ".join(separatori[:3]))
    if items and not any(clean(it.get("stanza")) for it in items):
        return True, ("Nessuno dei %d dispositivi ha una stanza: manca la "
                      "colonna Stanza." % len(items))
    return False, ""


LOCK_TIMEOUT = 20.0     # secondi di attesa prima di rinunciare
LOCK_STALE_AFTER = 120  # secondi dopo i quali un lock e' considerato abbandonato


class InventoryError(Exception):
    """Errore mostrabile all'utente."""


class LockBusy(InventoryError):
    pass


class BloccoIphoneNonSpedito(InventoryError):
    """Un iPhone si elimina solo dopo essere stato rispedito."""

    def __init__(self, item):
        self.item = item
        InventoryError.__init__(
            self,
            "%s non e' ancora stato rispedito al servizio telefonia\n"
            "e non puo' essere eliminato dall'inventario.\n\n"
            "Registra prima la spedizione con il pulsante Conferma spedizione,\n"
            "nel contenitore Iphone. Da quel momento restera' consultabile per\n"
            "%d mesi, e poi potra' essere eliminato."
            % (item["asset_tag"], MESI_CONSERVAZIONE))


class BloccoPrestitiAperti(InventoryError):
    """Un'operazione distruttiva non parte finche' ci sono prestiti aperti.

    Eliminare un dispositivo in prestito e' gia' vietato ovunque nel programma:
    e' fisicamente in mano a qualcuno, e toglierlo dall'inventario vuol dire
    perderne la traccia proprio mentre e' fuori. Un reset o una sostituzione lo
    toglierebbero comunque, in blocco e senza nemmeno passare dal cestino:
    quindi non partono. Prima si registrano i rientri.
    """

    QUANTI_NE_ELENCA = 15

    def __init__(self, prestati, che_cosa):
        self.prestati = list(prestati)
        righe = ["    %s  ->  %s" % (it.get("asset_tag") or it.get("imei"),
                                     it.get("prestato_a") or "")
                 for it in self.prestati[:self.QUANTI_NE_ELENCA]]
        if len(self.prestati) > self.QUANTI_NE_ELENCA:
            righe.append("    ...e altri %d"
                         % (len(self.prestati) - self.QUANTI_NE_ELENCA))
        InventoryError.__init__(
            self,
            "%s: ci sono %d dispositivi in prestito.\n\n%s\n\n"
            "Un dispositivo in prestito e' in mano a qualcuno: toglierlo\n"
            "dall'inventario vuol dire perderne la traccia proprio mentre e'\n"
            "fuori. Registra prima i rientri con il pulsante Registra rientro,\n"
            "nella stanza dove sono stati prestati.\n\n"
            "Non e' stato toccato niente."
            % (che_cosa, len(self.prestati), "\n".join(righe)))


class BloccoConservazione(InventoryError):
    """Un dispositivo spedito non si elimina prima dei tre mesi di conservazione."""

    def __init__(self, item, sblocco):
        self.item = item
        self.sblocco = sblocco
        InventoryError.__init__(
            self,
            "%s e' stato rispedito al servizio telefonia il %s.\n\n"
            "Va conservato in inventario per consultazione: potrai eliminarlo\n"
            "a partire dal %s."
            % (item["asset_tag"], item["spedito_il"], sblocco.strftime("%d/%m/%Y")))


class BloccoPrestito(InventoryError):
    """Un dispositivo in prestito non si sposta e non si elimina.

    Finche' e' nelle mani di una persona, l'inventario deve continuare a dire
    dov'e' e chi ce l'ha: spostarlo o cancellarlo perderebbe l'unica traccia
    che permette di andarlo a riprendere. Prima si registra il rientro.
    """

    def __init__(self, item):
        self.item = item
        InventoryError.__init__(
            self,
            T("%s e' in prestito a %s dal %s.\n\n"
              "Un dispositivo in prestito non si sposta e non si elimina:\n"
              "registra prima il rientro con il pulsante Registra rientro,\n"
              "nella stanza dove e' stato prestato.")
            % (item["asset_tag"], item.get("prestato_a", ""),
               item.get("prestato_il", "")))


# Quante copie automatiche si tengono nella cartella Backup. Servono a tornare
# indietro di qualche passo, non a fare da archivio storico: la copia che si
# conserva davvero e' quella che il tecnico si salva in locale.
COPIE_DA_TENERE = 10

# Come si chiamano i due file dentro la copia locale. I nomi sono fissi, non
# quelli del file di partenza: una copia salvata da una postazione deve poter
# essere riaperta da un'altra, dove l'inventario si chiama in un altro modo.
# Il cestino: quanti record ci stanno e per quanto. Non e' un archivio storico,
# e' la rete di sicurezza per l'eliminazione sbagliata di ieri.
ELIMINATI_MASSIMO = 300
ELIMINATI_GIORNI = 30
# Quanti dispositivi si possono eliminare in un colpo da un file Excel. Non e'
# un limite tecnico: e' il punto oltre il quale nessuno legge davvero l'elenco
# di quello che sta per sparire, e una conferma che non si legge non e' una
# conferma.
MASSIMO_ELIMINA_EXCEL = 200

NOME_DATI_NELLO_ZIP = "Inventario.xlsx"
NOME_IMPOSTAZIONI_NELLO_ZIP = "inventario_impostazioni.json"
NOME_ELIMINATI_NELLO_ZIP = "inventario_eliminati.json"


def e_una_copia_automatica(nome):
    """True se il nome e' quello che scrive copia_di_sicurezza().

    Si guarda la data nel nome, non l'estensione soltanto: nella cartella
    Backup puo' finire un file messo li' a mano, e quello non si tocca.
    """
    if nome.startswith("~$") or not nome.lower().endswith(".xlsx"):
        return False
    base = os.path.splitext(nome)[0]
    if base.endswith(")") and " (" in base:
        base = base[:base.rindex(" (")]       # "... (2)": stesso contenuto, ripetuto
    pezzi = base.rsplit("_", 2)
    if len(pezzi) < 3:
        return False
    try:
        datetime.strptime("_".join(pezzi[-2:]), "%Y-%m-%d_%H-%M-%S")
    except ValueError:
        return False
    return True


def current_user():
    try:
        user = getpass.getuser()
    except Exception:
        user = "sconosciuto"
    try:
        host = socket.gethostname()
    except Exception:
        host = platform.node() or "?"
    return "%s@%s" % (user, host)


def norm_tag(value):
    return " ".join(str(value or "").split()).upper()


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    return " ".join(str(value).split())


def is_iphone(tipo):
    return clean(tipo).lower() == TIPO_IPHONE


def _cestino_accanto(copia):
    """Il file del cestino che accompagna una copia di sicurezza."""
    return os.path.splitext(copia)[0] + "_eliminati.json"


def _data_eliminazione(voce):
    """La data di eliminazione di una voce del cestino, o None se illeggibile."""
    testo = clean((voce or {}).get("eliminato_il"))
    for formato in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(testo, formato)
        except ValueError:
            continue
    return None


def rinomine_in_elenco(vecchie, nuove):
    """Le voci rinominate fra due elenchi: coppie (vecchio, nuovo).

    Vale per le stanze e per i tipi di dispositivo: si modificano tutti e due in
    un riquadro di testo, una voce per riga. Rinominarne una vuol dire cambiare
    il testo di una riga, quindi la riga resta al suo posto. Si confrontano per
    posizione, ed e' una rinomina solo quando il nome di prima non c'e' piu' nel
    nuovo elenco e quello nuovo non c'era prima: cosi' aggiungere, togliere o
    riordinare non passa mai per una rinomina.

    Riordinare merita una parola: `[A, B, C]` che diventa `[C, B, A]` ha due
    posizioni cambiate, ma i nomi sono gli stessi. Se lo si prendesse per una
    rinomina si sposterebbero i dispositivi di due stanze intere, e nessuno
    l'aveva chiesto.
    """
    vecchie = [clean(v) for v in vecchie or []]
    nuove = [clean(n) for n in nuove or []]
    if set(vecchie) == set(nuove):
        return []                     # riordinate, non rinominate
    coppie = []
    for vecchio, nuovo in zip(vecchie, nuove):
        if vecchio == nuovo or not vecchio or not nuovo:
            continue
        if vecchio not in nuove and nuovo not in vecchie:
            coppie.append((vecchio, nuovo))
    return coppie


def rinomine_stanze(vecchie, nuove):
    """Le stanze rinominate. Vedi rinomine_in_elenco."""
    return rinomine_in_elenco(vecchie, nuove)


def rinomine_tipi(vecchi, nuovi):
    """I tipi di dispositivo rinominati. Vedi rinomine_in_elenco."""
    return rinomine_in_elenco(vecchi, nuovi)


def rinomina_tocca_gli_iphone(coppie):
    """Il tipo iPhone in una rinomina, da qualsiasi lato: (vecchio, nuovo) o None.

    "iPhone" non e' un'etichetta come le altre: e' la parola con cui il
    programma riconosce i telefoni. Da li' discendono l'IMEI al posto
    dell'asset tag, la stanza obbligata, la spedizione al servizio telefonia e
    il fatto che non si eliminino. Cambiarla vorrebbe dire che da quel momento
    i telefoni non sono piu' telefoni, senza che nessuno se ne accorga.
    """
    for vecchio, nuovo in coppie:
        if is_iphone(vecchio) or is_iphone(nuovo):
            return (vecchio, nuovo)
    return None


def new_item(asset_tag="", tipo="", modello="", seriale="", stanza="", note="",
             prestato_a="", prestato_il="", imei="", restituito_da="", stato="",
             spedito_il=""):
    item = {
        "asset_tag": norm_tag(asset_tag),
        "tipo": clean(tipo),
        "modello": clean(modello),
        "seriale": clean(seriale),
        "imei": clean(imei),
        "restituito_da": clean(restituito_da),
        "stanza": clean(stanza),
        "stato": clean(stato),
        "prestato_a": clean(prestato_a),
        "prestato_il": clean(prestato_il),
        "spedito_il": clean(spedito_il),
        "note": clean(note),
        "modificato_il": "",
        "modificato_da": "",
    }
    return normalize_state(normalize_identity(item))


def normalize_identity(item):
    """Per un iPhone l'IMEI e' l'identificativo: se manca l'asset tag, lo sostituisce."""
    if not item.get("asset_tag") and item.get("imei"):
        item["asset_tag"] = norm_tag(item["imei"])
    return item


def _stato_canonico(valore, ammessi):
    """Riporta all'italiano uno stato scritto in inglese in un file importato."""
    testo = clean(valore)
    if not testo or testo in ammessi:
        return testo
    from .lingua import STATI_EN
    rovescio = dict((v.lower(), k) for k, v in STATI_EN.items())
    return rovescio.get(testo.lower(), testo)


def normalize_iphone(item):
    """Un iPhone ha l'IMEI e basta: niente asset tag, seriale o prestiti.

    L'asset tag resta valorizzato in memoria come chiave interna - serve a
    identificare la riga in ogni operazione - ma vale sempre l'IMEI e non viene
    mai mostrato ne' scritto sul file.
    """
    if is_iphone(item.get("tipo")):
        item["seriale"] = ""
        item["prestato_a"] = ""
        item["prestato_il"] = ""
        if item.get("imei"):
            item["asset_tag"] = norm_tag(item["imei"])
    return item


def valore_visibile(item, field):
    """Il valore da mostrare o scrivere: per un iPhone l'asset tag resta vuoto."""
    if field == "asset_tag" and is_iphone(item.get("tipo")):
        return ""
    return item.get(field, "")


def normalize_state(item, stati=None):
    """Mette a posto lo stato.

    Due casi sono automatici e vincono sempre: un iPhone e' "Da Rispedire", un
    dispositivo prestato e' "In prestito". Negli altri casi si tiene lo stato
    scelto dall'utente, purche' sia fra quelli previsti.
    """
    ammessi = list(stati or STATI)
    normalize_iphone(item)
    item["stato"] = _stato_canonico(item.get("stato"), ammessi)
    if is_iphone(item.get("tipo")):
        item["stato"] = SPEDITO if item.get("spedito_il") else DA_RISPEDIRE
    elif item.get("prestato_a"):
        item["stato"] = NON_DISPONIBILE
    elif clean(item.get("stato")) not in ammessi:
        item["stato"] = DISPONIBILE
    return item


def is_on_loan(item):
    return bool(item.get("prestato_a"))


def is_shipped(item):
    return bool(item.get("spedito_il"))


def _somma_mesi(quando, mesi):
    anno = quando.year + (quando.month - 1 + mesi) // 12
    mese = (quando.month - 1 + mesi) % 12 + 1
    giorno = min(quando.day, calendar.monthrange(anno, mese)[1])
    return quando.replace(year=anno, month=mese, day=giorno)


def eliminabile_dal(item):
    """Data dalla quale un dispositivo spedito puo' essere eliminato, o None."""
    quando = clean(item.get("spedito_il"))
    if not quando:
        return None
    for formato in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return _somma_mesi(datetime.strptime(quando, formato), MESI_CONSERVAZIONE)
        except ValueError:
            continue
    return None


def prestiti_aperti(items, stanza=None):
    """I dispositivi con un prestito aperto, in tutto l'inventario o in una stanza.

    Serve a chi sta per cancellarne tanti in un colpo solo - reset,
    sostituzione - per fermarsi prima invece di accorgersene dopo.
    """
    return [it for it in items
            if is_on_loan(it)
            and (stanza is None or clean(it.get("stanza")) == clean(stanza))]


def puo_essere_eliminato(item, adesso=None):
    """(True/False, data di sblocco).

    Un iPhone non ancora rispedito non si elimina mai: va prima registrata la
    spedizione, e da quel momento restano tre mesi di conservazione. In quel
    caso la data di sblocco e' None, perche' non e' ancora determinabile.
    """
    if is_iphone(item.get("tipo")) and not is_shipped(item):
        return False, None
    sblocco = eliminabile_dal(item)
    if sblocco is None:
        return True, None
    return (adesso or datetime.now()) >= sblocco, sblocco


def testo_spedizione(item):
    """La frase da mostrare accanto a un dispositivo spedito."""
    sblocco = eliminabile_dal(item)
    if sblocco is None:
        return ""
    return TESTO_SPEDIZIONE % (item["spedito_il"], sblocco.strftime("%d/%m/%Y"))


class _Lock(object):
    """Lock esclusivo basato su un file accanto ai dati."""

    def __init__(self, data_path):
        self.path = os.path.join(
            os.path.dirname(os.path.abspath(data_path)),
            "." + os.path.basename(data_path) + ".lock",
        )
        self.fd = None

    def _holder(self):
        try:
            with open(self.path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except (OSError, ValueError):
            return {}

    def __enter__(self):
        deadline = time.time() + LOCK_TIMEOUT
        while True:
            try:
                self.fd = os.open(self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                info = json.dumps(
                    {"utente": current_user(), "pid": os.getpid(), "ts": time.time()}
                )
                os.write(self.fd, info.encode("utf-8"))
                os.close(self.fd)
                self.fd = None
                return self
            except FileExistsError:
                holder = self._holder()
                # L'eta' si misura sul file, non sul contenuto: subito dopo la
                # creazione il file e' ancora vuoto e non va scambiato per
                # abbandonato.
                try:
                    age = time.time() - os.stat(self.path).st_mtime
                except OSError:
                    continue
                if age > LOCK_STALE_AFTER:
                    # Lock abbandonato (PC spento, crash): lo rimuoviamo.
                    try:
                        os.remove(self.path)
                    except OSError:
                        pass
                    continue
                if time.time() >= deadline:
                    raise LockBusy(
                        "L'inventario e' in uso da %s. Riprova tra qualche istante."
                        % (holder.get("utente") or "un altro utente")
                    )
                time.sleep(0.4)
            except OSError as exc:
                raise InventoryError(
                    T("Impossibile accedere alla cartella di rete:\n%s") % exc
                )

    def __exit__(self, *exc):
        try:
            os.remove(self.path)
        except OSError:
            pass
        return False


class InventoryStore(object):
    def __init__(self, data_path, iphone_room=None, stati=None):
        self.path = os.path.abspath(data_path)
        # Stanza in cui gli iPhone stanno per forza; None disattiva il vincolo.
        self.iphone_room = iphone_room
        self.stati = list(stati or STATI)
        # Record toccati da quando il programma e' aperto: serve a ricordare di
        # tenersi una copia locale, non a fare statistiche.
        self.modifiche = 0
        self.items = []
        self._stamp = None
        # Le copie automatiche buttate via dall'ultima rotazione: chi vuole puo'
        # dirlo, ma nessuno deve dipenderne.
        self.copie_scartate = []
        # Se il cestino e' entrato nell'ultima copia locale, e quanti record ne
        # sono tornati con l'ultimo ripristino.
        self.eliminati_nella_copia = False
        self.eliminati_ripristinati = 0
        # Voci del cestino tolte dall'ultimo ripristino perche' il dispositivo
        # era gia' in inventario.
        self.tolti_perche_presenti = []
        # Le stanze dell'inventario. None significa "non me le hanno dette":
        # in quel caso una stanza vale purche' ci sia. L'interfaccia le passa
        # sempre, ed e' li' che serve il controllo.
        self.stanze = None
        # Dispositivi che l'ultima eliminazione ha cancellato per sempre invece
        # di metterli nel cestino, perche' non ci stavano.
        self.cancellati_per_sempre = []
        # Voci tolte dal cestino dall'ultimo ripristino, perche' il dispositivo
        # e' tornato in inventario.
        self.tolti_dal_ripristino = []

    def stanza_canonica(self, valore):
        """Il nome esatto della stanza, o "" se quel nome non e' una stanza.

        Un nome che non c'e' - "Cantina" - non e' una stanza: il dispositivo
        non comparirebbe in nessuna scheda, non uscirebbe da nessuna
        esportazione per stanza, e per ritrovarlo bisognerebbe gia' sapere che
        c'e'. Vale come se non ne avesse nessuna, e l'importazione la chiede.

        Quello che invece si perdona e' il modo di scriverla: lo spazio di
        troppo e la maiuscola sbagliata - "digital  kiosk" e' Digital Kiosk -
        perche' quella e' la stanza giusta scritta male, non un'altra stanza.
        Il nome che entra in inventario e' comunque quello ufficiale, o si
        ritroverebbero due stanze dove ce n'e' una.
        """
        nome = clean(valore)
        if not nome:
            return ""
        if self.stanze is None:
            return nome
        for stanza in list(self.stanze) + [self.iphone_room]:
            if clean(stanza) and clean(stanza).lower() == nome.lower():
                return clean(stanza)
        return ""

    def stanza_ammessa(self, valore):
        """Vero se quel nome e' davvero una stanza dell'inventario."""
        return bool(self.stanza_canonica(valore))

    def _enforce_iphone_room(self, item):
        """Un iPhone appartiene sempre alla sua stanza, comunque lo si registri."""
        if self.iphone_room and is_iphone(item.get("tipo")):
            item["stanza"] = self.iphone_room
        return item

    # ------------------------------------------------------------ lettura

    def exists(self):
        return os.path.exists(self.path)

    def create_if_missing(self):
        """Crea l'inventario, e accanto le impostazioni con i valori di partenza.

        Scrivere subito il file delle impostazioni serve a due cose: rende
        visibile com'e' configurato l'inventario, e lo rende uguale per tutti i
        tecnici invece di dipendere dai valori predefiniti del programma.
        """
        if self.exists():
            return False
        folder = os.path.dirname(self.path)
        if folder and not os.path.isdir(folder):
            raise InventoryError(T("La cartella %s non esiste.") % folder)
        self._write([])
        from . import config
        if not os.path.exists(config.shared_config_path(self.path)):
            try:
                config.save_shared_config(self.path, config.load_shared_config(self.path))
            except OSError:
                pass          # cartella in sola lettura: si va avanti con i predefiniti
        return True

    def _disk_stamp(self):
        try:
            st = os.stat(self.path)
        except OSError:
            return None
        return (st.st_mtime, st.st_size)

    def changed_on_disk(self):
        """True se qualcun altro ha scritto il file dopo la nostra lettura."""
        return self._disk_stamp() != self._stamp

    def load(self):
        self.items = self._read()
        self._stamp = self._disk_stamp()
        return self.items

    def _read(self):
        if not self.exists():
            return []
        try:
            wb = load_workbook(self.path, read_only=True, data_only=True)
        except Exception as exc:
            raise InventoryError(T("Impossibile leggere %s:\n%s") % (self.path, exc))
        try:
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                return []
            mapping = map_headers(header)
            items = []
            for row in rows:
                item = _row_to_item(row, mapping)
                if item:
                    items.append(item)
            return items
        finally:
            wb.close()

    # ---------------------------------------------------------- scrittura

    def _write(self, items):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append([HEADERS[f] for f in ALL_FIELDS])
        for item in items:
            ws.append([valore_visibile(item, f) for f in ALL_FIELDS])
        _style_sheet(ws, len(items))
        tmp = "%s.tmp-%d-%s" % (self.path, os.getpid(), uuid.uuid4().hex[:8])
        try:
            wb.save(tmp)
            os.replace(tmp, self.path)
        except Exception as exc:
            try:
                os.remove(tmp)
            except OSError:
                pass
            raise InventoryError(T("Impossibile salvare %s:\n%s") % (self.path, exc))
        finally:
            wb.close()

    def _apply(self, operation):
        """Esegue `operation(items)` su dati freschi, dentro il lock, e salva.

        `operation` puo' sollevare InventoryError per annullare tutto.
        Ritorna il valore restituito da `operation`.
        """
        with _Lock(self.path):
            items = self._read()
            prima = dict((it["asset_tag"], dict(it)) for it in items)
            result = operation(items)
            items.sort(key=lambda it: (it.get("stanza", ""), it.get("asset_tag", "")))
            self._write(items)
            self.items = items
            self._stamp = self._disk_stamp()
            self.modifiche += _quanti_cambiati(prima, items)
        return result

    # ---------------------------------------------------------- operazioni

    def add(self, item):
        item = dict(item)
        item["asset_tag"] = norm_tag(item.get("asset_tag"))
        if not item["asset_tag"]:
            raise InventoryError(T("L'asset tag e' obbligatorio."))

        def op(items):
            gia_presente = next((it for it in items
                                 if it["asset_tag"] == item["asset_tag"]), None)
            if gia_presente is not None:
                # dire solo "esiste gia'" costringe a cercarlo a mano: si dice
                # dov'e' e che cos'e', cosi' si capisce subito se e' un errore
                # di battitura o il dispositivo che si aveva gia' registrato
                dettagli = [d for d in (gia_presente.get("modello"),
                                        gia_presente.get("stanza")) if d]
                raise InventoryError(
                    T("%s e' gia' in inventario%s.\n\n"
                    "Non e' stato inserito niente: due dispositivi non possono "
                    "avere lo stesso identificativo.\n\n"
                    "Se e' un dispositivo diverso, controlla il codice; se e' "
                    "lo stesso, modificalo invece di reinserirlo.")
                    % (item["asset_tag"],
                       "  -  " + ", ".join(dettagli) if dettagli else ""))
            self._enforce_iphone_room(item)
            self._stanza_o_rifiuta(item)
            normalize_state(item, self.stati)
            _stamp_item(item)
            items.append(item)
            return item["asset_tag"]      # cosi' chi chiama sa che e' andata

        return self._apply(op)

    def update(self, old_tag, item):
        old_tag = norm_tag(old_tag)
        item = dict(item)
        item["asset_tag"] = norm_tag(item.get("asset_tag"))
        if not item["asset_tag"]:
            raise InventoryError(T("L'asset tag e' obbligatorio."))

        def op(items):
            index = _index_of(items, old_tag)
            if index is None:
                raise InventoryError(
                    T("L'articolo %s non esiste piu': e' stato eliminato da un altro utente.")
                    % old_tag
                )
            if item["asset_tag"] != old_tag and _index_of(items, item["asset_tag"]) is not None:
                raise InventoryError(
                    T("L'asset tag %s e' gia' presente nell'inventario.") % item["asset_tag"]
                )
            precedente = items[index]
            if is_on_loan(precedente) and clean(item.get("stanza")) != clean(
                    precedente.get("stanza")):
                raise BloccoPrestito(precedente)
            self._enforce_iphone_room(item)
            self._stanza_o_rifiuta(item)
            normalize_state(item, self.stati)
            _stamp_item(item)
            items[index] = item

        return self._apply(op)

    def _stanza_o_rifiuta(self, item):
        """Scrive nel dispositivo il nome ufficiale della sua stanza.

        Un dispositivo in una stanza che non esiste non comparirebbe in nessuna
        scheda: e' lo stesso di non averne una. Dalla finestra non puo'
        succedere - la stanza si sceglie da una tendina - ma la regola vale per
        chiunque scriva nell'archivio, non solo per chi passa di li'.
        """
        nome = self.stanza_canonica(item.get("stanza"))
        if not nome:
            raise InventoryError(
                T("%s non e' una stanza dell'inventario.\n\n"
                "Un dispositivo senza una stanza vera non comparirebbe in "
                "nessuna scheda. Scegli una delle stanze esistenti, o creala "
                "prima dalle impostazioni.")
                % (clean(item.get("stanza")) or T("(nessuna stanza)")))
        item["stanza"] = nome

    def eccesso_cestino(self, quanti):
        """Quanti dei prossimi `quanti` eliminati non entrerebbero nel cestino.

        Il cestino tiene ELIMINATI_MASSIMO record: oltre quel numero qualcosa
        deve sparire per sempre, e chi elimina deve poter decidere che cosa
        invece di scoprirlo dopo.
        """
        dentro = len(self._pota_eliminati(self._leggi_eliminati()))
        return max(0, dentro + quanti - ELIMINATI_MASSIMO)

    def delete(self, tags, in_eccesso="cestino"):
        """Elimina i dispositivi, salvo quelli spediti da meno di tre mesi.

        in_eccesso dice che fare di quelli che nel cestino non ci stanno:
        "cestino" li mette dentro al posto dei piu' vecchi, che escono per
        sempre; "definitivo" li cancella subito senza passare dal cestino, che
        resta com'e'. Quelli cancellati per sempre restano in
        `cancellati_per_sempre`, per poterlo dire a chi ha eliminato.
        """
        wanted = set(norm_tag(t) for t in tags)

        def op(items):
            for it in items:
                if it["asset_tag"] not in wanted:
                    continue
                if is_on_loan(it):
                    raise BloccoPrestito(it)
                libero, sblocco = puo_essere_eliminato(it)
                if libero:
                    continue
                if sblocco is None:
                    raise BloccoIphoneNonSpedito(it)
                raise BloccoConservazione(it, sblocco)
            tolti[:] = [dict(it) for it in items if it["asset_tag"] in wanted]
            before = len(items)
            items[:] = [it for it in items if it["asset_tag"] not in wanted]
            return before - len(items)

        tolti = []
        quanti = self._apply(op)
        self.cancellati_per_sempre = []
        if in_eccesso == "definitivo":
            # non ci stanno tutti: chi e' in fondo alla fila non entra affatto
            eccesso = self.eccesso_cestino(len(tolti))
            if eccesso:
                taglio = max(0, len(tolti) - eccesso)
                self.cancellati_per_sempre = tolti[taglio:]
                tolti = tolti[:taglio]
        # Il cestino si aggiorna dopo: se l'eliminazione fallisce non ci finisce
        # dentro niente, e se fallisce il cestino l'eliminazione resta valida.
        self.aggiungi_agli_eliminati(tolti)
        return quanti

    # ------------------------------------------------- eliminati di recente

    def _percorso_eliminati(self):
        from . import config
        return config.deleted_path(self.path)

    def _leggi_eliminati(self):
        try:
            with open(self._percorso_eliminati(), "r", encoding="utf-8") as fh:
                voci = json.load(fh)
        except (OSError, ValueError):
            return []
        return voci if isinstance(voci, list) else []

    def _scrivi_eliminati(self, voci):
        percorso = self._percorso_eliminati()
        tmp = percorso + ".tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as fh:
                json.dump(voci, fh, indent=2, ensure_ascii=False)
            os.replace(tmp, percorso)
        except OSError:
            pass          # il cestino e' un di piu': non fa fallire un'eliminazione

    def _pota_eliminati(self, voci, adesso=None):
        """Toglie quelli scaduti e tiene solo gli ultimi ELIMINATI_MASSIMO."""
        adesso = adesso or datetime.now()
        limite = adesso - timedelta(days=ELIMINATI_GIORNI)
        vivi = []
        for voce in voci:
            quando = _data_eliminazione(voce)
            if quando is not None and quando < limite:
                continue
            vivi.append(voce)
        vivi.sort(key=lambda v: _data_eliminazione(v) or datetime.min, reverse=True)
        return vivi[:ELIMINATI_MASSIMO]

    def aggiungi_agli_eliminati(self, items, orfani=False):
        """Mette nel cestino i dispositivi appena tolti dall'inventario.

        Si tiene la scheda intera, non il solo identificativo: il ripristino
        deve rimettere il dispositivo com'era, stanza compresa. Per un orfano -
        un dispositivo rimasto senza stanza perche' la stanza e' stata tolta -
        la stanza di partenza non c'e' piu', e al ripristino verra' chiesta.
        """
        if not items:
            return []
        quando = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        chi = current_user()
        voci = self._leggi_eliminati()
        nuove = []
        for item in items:
            scheda = dict(item)
            if orfani:
                # La stanza che aveva non esiste piu': tenerla nella scheda
                # rimetterebbe il dispositivo in una stanza fantasma. Il nome
                # resta scritto nella voce, che serve a ritrovarlo.
                scheda["stanza"] = ""
            nuove.append({"scheda": scheda,
                          "asset_tag": norm_tag(scheda.get("asset_tag")),
                          "tipo": clean(scheda.get("tipo")),
                          "stanza": clean(item.get("stanza")),
                          "orfano": bool(orfani),
                          "eliminato_il": quando,
                          "eliminato_da": chi})
        voci = self._pota_eliminati(nuove + voci)
        self._scrivi_eliminati(voci)
        return nuove

    def eliminati(self, cerca=""):
        """Il cestino, dal piu' recente. La potatura avviene qui, alla lettura.

        Cosi' un record scaduto non si vede mai, anche se nessuno ha eliminato
        niente per settimane e il file e' rimasto fermo.
        """
        voci = self._pota_eliminati(self._leggi_eliminati())
        cerca = clean(cerca).lower()
        if not cerca:
            return voci
        return [v for v in voci
                if cerca in ("%s %s %s" % (v.get("asset_tag", ""), v.get("tipo", ""),
                                           v.get("stanza", ""))).lower()]

    def togli_dal_cestino(self, tags=None):
        """Toglie dal cestino i record che intanto sono (tornati) in inventario.

        Un dispositivo non puo' essere insieme in elenco e fra gli eliminati di
        recente: sarebbe un doppione fra due posti che si contraddicono, e chi
        lo ripescasse dal cestino ne creerebbe una seconda copia. Ogni strada
        che rimette dentro un identificativo - importazione, inserimento
        singolo, controllo doppioni - passa di qui.

        `tags` limita la pulizia a quegli identificativi; senza, guarda tutto
        l'inventario. Si toglie solo cio' che in inventario c'e' davvero: il
        cestino non si svuota per sbaglio.

        Ritorna la lista dei tolti, con la stanza in cui il dispositivo si trova
        adesso in inventario - la domanda che si fa chi legge il riepilogo.
        """
        voci = self._pota_eliminati(self._leggi_eliminati())
        if not voci:
            return []
        in_inventario = dict((norm_tag(i.get("asset_tag")), i) for i in self.items)
        volute = set(norm_tag(t) for t in tags) if tags is not None else None
        tolti, restano = [], []
        for voce in voci:
            tag = norm_tag(voce.get("asset_tag"))
            dentro = in_inventario.get(tag)
            if dentro is None or (volute is not None and tag not in volute):
                restano.append(voce)
                continue
            tolti.append({"asset_tag": tag,
                          "tipo": clean(voce.get("tipo")),
                          "stanza": clean(dentro.get("stanza")) or SENZA_STANZA})
        if tolti:
            self._scrivi_eliminati(restano)
        return tolti

    def stanza_del_ritorno(self, voce):
        """La stanza in cui un record del cestino tornerebbe, o "" se va chiesta.

        Non basta che la voce porti scritto un nome di stanza: quella stanza
        puo' non esserci piu' - tolta, o rinominata mentre il record stava nel
        cestino. Rimetterci dentro il dispositivo lo farebbe sparire da tutte
        le schede, che e' lo stesso di non avere stanza. Allora si chiede,
        esattamente come per un orfano.
        """
        if voce.get("orfano"):
            return ""
        scheda = voce.get("scheda") or {}
        return self.stanza_canonica(clean(scheda.get("stanza"))
                                    or clean(voce.get("stanza")))

    def ripristina_eliminati(self, chiavi, stanza=None):
        """Rimette in inventario i record scelti, nella stanza che avevano.

        `chiavi` sono gli asset tag come stanno nel cestino. Un orfano non ha
        una stanza a cui tornare: per quello si passa `stanza`.

        Ritorna (ripristinati, saltati). `ripristinati` sono dizionari con
        asset tag, tipo e la stanza in cui il dispositivo e' tornato: chi ha
        chiesto il ripristino deve poter leggere dove sono finiti, non solo
        quanti erano. `saltati` sono coppie (asset tag, motivo).

        Un identificativo che intanto e' tornato in inventario per un'altra
        strada non e' un errore da segnalare: e' una voce del cestino che non
        ha piu' ragione di esistere. Viene tolta, e finisce in
        `tolti_perche_presenti` con la stanza in cui il dispositivo si trova
        adesso - la prima cosa che si vuole sapere.
        """
        volute = [norm_tag(c) for c in chiavi]
        voci = self._pota_eliminati(self._leggi_eliminati())
        per_tag = dict((v.get("asset_tag"), v) for v in voci)
        da_rimettere, saltati = [], []
        for tag in volute:
            voce = per_tag.get(tag)
            if voce is None:
                saltati.append((tag, T("non e' piu' fra gli eliminati di recente")))
                continue
            scheda = dict(voce.get("scheda") or {})
            dove = self.stanza_del_ritorno(voce)
            dove = dove or self.stanza_canonica(stanza) or clean(stanza)
            if not dove:
                saltati.append((tag, T("non aveva una stanza: indica dove rimetterlo")))
                continue
            scheda["stanza"] = dove
            da_rimettere.append((tag, scheda))

        gia_dentro = []

        def op(items):
            presenti = dict((norm_tag(i.get("asset_tag")), i) for i in items)
            rimessi = []
            for tag, scheda in da_rimettere:
                if tag in presenti:
                    gia_dentro.append(
                        {"asset_tag": tag,
                         "tipo": clean(scheda.get("tipo")),
                         "stanza": clean(presenti[tag].get("stanza")) or SENZA_STANZA})
                    continue
                self._enforce_iphone_room(scheda)
                normalize_state(scheda, self.stati)
                _stamp_item(scheda)
                items.append(scheda)
                presenti[tag] = scheda
                rimessi.append({"asset_tag": tag,
                                "tipo": clean(scheda.get("tipo")),
                                "stanza": clean(scheda.get("stanza"))})
            return rimessi

        rimessi = self._apply(op) if da_rimettere else []
        self.tolti_perche_presenti = gia_dentro
        # Escono dal cestino sia quelli tornati dentro adesso, sia quelli che
        # c'erano gia': in tutti e due i casi il dispositivo e' in inventario,
        # e restare anche nel cestino non avrebbe senso.
        tolti = set(r["asset_tag"] for r in rimessi) | set(
            r["asset_tag"] for r in gia_dentro)
        if tolti:
            self._scrivi_eliminati([v for v in voci if v.get("asset_tag") not in tolti])
        return rimessi, saltati

    def trasloca_stanza(self, vecchia, nuova):
        """Porta in un'altra stanza tutto quello che sta in questa.

        Serve quando la stanza sparisce ma il suo ruolo no: e' la stanza degli
        iPhone, o quella dove si prestano i dispositivi. I telefoni devono
        stare da qualche parte, e i prestiti aperti pure.

        A differenza di Sposta in stanza, qui **si spostano anche i dispositivi
        in prestito**: la regola che li tiene fermi serve a non perderne la
        traccia, e lasciarli in una stanza che non esiste piu' sarebbe
        esattamente perderla. Il prestito resta aperto: cambia dove il
        dispositivo risulta registrato, non chi ce l'ha.

        Ritorna {"totale": n, "prestiti": n, "iphone": n}.
        """
        vecchia, nuova = clean(vecchia), clean(nuova)
        if not vecchia or not nuova or vecchia == nuova:
            return {"totale": 0, "prestiti": 0, "iphone": 0}

        def op(items):
            conto = {"totale": 0, "prestiti": 0, "iphone": 0}
            for it in items:
                if clean(it.get("stanza")) != vecchia:
                    continue
                it["stanza"] = nuova
                _stamp_item(it)
                conto["totale"] += 1
                if is_on_loan(it):
                    conto["prestiti"] += 1
                if is_iphone(it.get("tipo")):
                    conto["iphone"] += 1
            return conto

        return self._apply(op)

    def porta_via_gli_orfani(self, stanze):
        """Toglie dall'inventario i dispositivi delle stanze indicate.

        Serve quando una stanza viene tolta dalle impostazioni: i dispositivi
        che ci stavano non hanno piu' un posto, e restare in elenco con il nome
        di una stanza che non esiste e' peggio che sparire - non si trovano con
        i filtri e non si sa che fine hanno fatto. Vanno nel cestino segnati
        come orfani, e al ripristino si chiedera' in che stanza rimetterli.

        Ritorna la lista dei dispositivi portati via.
        """
        volute = set(clean(s) for s in stanze if clean(s))
        if not volute:
            return []
        portati = []

        def op(items):
            # Un dispositivo in prestito non si elimina, e finire nel cestino
            # sarebbe eliminarlo. L'unica strada per portarlo altrove e' il
            # trasloco della stanza, che glielo cambia sotto invece di
            # toglierlo dall'inventario. Vale identico per un iPhone non ancora
            # eliminabile: dalla finestra non ci si arriva - togliere la stanza
            # degli iPhone fa scattare il trasloco - ma la regola sta qui, non
            # nella finestra che per caso ci passa davanti.
            for it in items:
                if clean(it.get("stanza")) not in volute:
                    continue
                if is_on_loan(it):
                    raise BloccoPrestito(it)
                libero, sblocco = puo_essere_eliminato(it)
                if not libero:
                    if sblocco is None:
                        raise BloccoIphoneNonSpedito(it)
                    raise BloccoConservazione(it, sblocco)
            portati[:] = [dict(it) for it in items if clean(it.get("stanza")) in volute]
            items[:] = [it for it in items if clean(it.get("stanza")) not in volute]
            return len(portati)

        self._apply(op)
        self.aggiungi_agli_eliminati(portati, orfani=True)
        return portati

    def stanze_con_prestiti_aperti(self, stanze):
        """Fra quelle indicate, le stanze che tengono ancora un prestito aperto.

        Una stanza cosi' non si puo' svuotare nel cestino: quei dispositivi sono
        nelle mani di qualcuno. Se sparisce va traslocata, come quella degli
        iPhone.
        """
        aperte = []
        for stanza in stanze:
            stanza = clean(stanza)
            if stanza and any(clean(i.get("stanza")) == stanza and is_on_loan(i)
                              for i in self.items):
                aperte.append(stanza)
        return aperte

    def quanti_nelle_stanze(self, stanze):
        """Quanti dispositivi ci sono in ognuna delle stanze indicate."""
        conteggio = {}
        for stanza in stanze:
            stanza = clean(stanza)
            if not stanza:
                continue
            conteggio[stanza] = sum(1 for i in self.items
                                    if clean(i.get("stanza")) == stanza)
        return conteggio

    def copia_di_sicurezza(self):
        """Salva il file dati nella cartella Backup, prima di un'operazione grossa.

        Il nome porta la data del file salvato, non quella della copia: due
        reset di fila sullo stesso inventario non producono due file uguali, e
        cercando una versione si guarda a quando risale il contenuto.

        Accanto alla copia viene salvato anche il cestino di quel momento, come
        `<nome>_eliminati.json`: l'inventario e gli eliminati di recente sono
        una cosa sola, e tornare indietro sull'uno senza l'altro lascerebbe nel
        cestino roba che nel frattempo e' rientrata in inventario.
        """
        from . import config

        cartella = config.backup_dir()
        if cartella is None:
            raise InventoryError(
                T("Non riesco a creare la cartella delle copie di sicurezza.\n\n"
                "L'operazione e' stata annullata: nessun dato e' stato toccato."))
        try:
            quando = datetime.fromtimestamp(os.path.getmtime(self.path))
        except OSError:
            quando = datetime.now()
        radice = os.path.splitext(os.path.basename(self.path))[0]
        base = "%s_%s" % (radice, quando.strftime("%Y-%m-%d_%H-%M-%S"))
        destinazione = os.path.join(cartella, base + ".xlsx")
        contatore = 2
        while os.path.exists(destinazione):
            destinazione = os.path.join(cartella, "%s (%d).xlsx" % (base, contatore))
            contatore += 1
        try:
            shutil.copy2(self.path, destinazione)
        except OSError as exc:
            raise InventoryError(
                T("Non riesco a creare la copia di sicurezza:\n%s\n\n"
                "L'operazione e' stata annullata: nessun dato e' stato toccato.") % exc)
        # Il cestino dopo i dati e senza far fallire niente: e' un di piu' utile,
        # non una condizione perche' l'operazione possa procedere. Si scrive
        # anche quando e' vuoto: senza il gemello, il ripristino non saprebbe
        # che in quel momento il cestino era vuoto e lascerebbe dentro quello di
        # adesso - con dentro dispositivi che il ripristino ha appena rimesso in
        # inventario.
        cestino = self._percorso_eliminati()
        try:
            if os.path.exists(cestino):
                shutil.copy2(cestino, _cestino_accanto(destinazione))
            else:
                with open(_cestino_accanto(destinazione), "w",
                          encoding="utf-8") as fh:
                    json.dump([], fh)
        except OSError:
            pass
        self.copie_scartate = self._tieni_solo_le_ultime(cartella)
        return destinazione

    def _tieni_solo_le_ultime(self, cartella, quante=None):
        """Cancella le copie automatiche piu' vecchie oltre le ultime `quante`.

        Va chiamata dopo aver scritto la copia nuova, mai prima: se la scrittura
        fallisse si sarebbe fatto spazio buttando via una copia buona.

        Una copia che non si riesce a cancellare - aperta in Excel da un collega,
        permessi negati sulla share - non ferma niente: la copia nuova c'e' gia',
        e l'operazione che l'ha chiesta deve andare avanti.
        """
        quante = COPIE_DA_TENERE if quante is None else quante
        try:
            nomi = os.listdir(cartella)
        except OSError:
            return []
        copie = []
        for nome in nomi:
            if not e_una_copia_automatica(nome):
                continue
            percorso = os.path.join(cartella, nome)
            try:
                copie.append((os.path.getmtime(percorso), nome, percorso))
            except OSError:
                continue
        copie.sort(reverse=True)          # dalla piu' recente, come Ripristina
        scartate = []
        for _quando, _nome, percorso in copie[quante:]:
            # il cestino gemello se ne va con la sua copia, o resterebbe li' per
            # sempre a occupare posto senza piu' un inventario a cui riferirsi
            gemello = _cestino_accanto(percorso)
            if os.path.exists(gemello):
                try:
                    os.remove(gemello)
                except OSError:
                    pass
            try:
                os.remove(percorso)
                scartate.append(percorso)
            except OSError:
                pass
        return scartate

    def copia_in(self, destinazione):
        """Copia l'inventario, com'e' in questo istante, nel percorso indicato.

        La copia si prende dentro il lock: se in quel momento un altro tecnico
        sta salvando, si aspetta che abbia finito. Cosi' il file portato via non
        e' mai un inventario colto a meta' scrittura, ed e' aggiornato al
        secondo in cui viene chiesto.

        Accanto all'inventario viene salvato anche il file delle impostazioni -
        stanze, tipi, stati - perche' da soli i dati non bastano a ricostruire
        l'inventario com'era, e quello degli eliminati di recente: se la copia
        serve dopo che la share e' sparita, sparirebbe con lei anche il cestino,
        cioe' l'unico posto da cui si ripesca quello tolto per sbaglio.

        Ritorna (file dati, file impostazioni o None, quanti dispositivi).
        `eliminati_nella_copia` dice se il cestino ci e' entrato.
        """
        from . import config

        cartella = os.path.dirname(os.path.abspath(destinazione))
        if not os.path.isdir(cartella):
            raise InventoryError(T("La cartella non esiste:\n%s") % cartella)
        sorgente = config.shared_config_path(self.path)
        cestino = self._percorso_eliminati()
        self.eliminati_nella_copia = False
        if destinazione.lower().endswith(".zip"):
            with _Lock(self.path):
                if not os.path.exists(self.path):
                    raise InventoryError(T("L'inventario non c'e' piu':\n%s") % self.path)
                quanti = len(self._read())
                impostazioni = None
                try:
                    with zipfile.ZipFile(destinazione, "w",
                                         zipfile.ZIP_DEFLATED) as archivio:
                        archivio.write(self.path, NOME_DATI_NELLO_ZIP)
                        if os.path.exists(sorgente):
                            archivio.write(sorgente, NOME_IMPOSTAZIONI_NELLO_ZIP)
                            impostazioni = NOME_IMPOSTAZIONI_NELLO_ZIP
                        # anche vuoto: senza, il ripristino non saprebbe che
                        # in quel momento il cestino era vuoto
                        if os.path.exists(cestino):
                            archivio.write(cestino, NOME_ELIMINATI_NELLO_ZIP)
                        else:
                            archivio.writestr(NOME_ELIMINATI_NELLO_ZIP, "[]")
                        self.eliminati_nella_copia = True
                except OSError as exc:
                    raise InventoryError(T("Non riesco a salvare la copia:\n%s") % exc)
            return destinazione, impostazioni, quanti

        with _Lock(self.path):
            if not os.path.exists(self.path):
                raise InventoryError(T("L'inventario non c'e' piu':\n%s") % self.path)
            try:
                shutil.copy2(self.path, destinazione)
            except OSError as exc:
                raise InventoryError(T("Non riesco a salvare la copia:\n%s") % exc)
            quanti = len(self._read())
            impostazioni = None
            if os.path.exists(sorgente):
                accanto = os.path.splitext(destinazione)[0] + "_impostazioni.json"
                try:
                    shutil.copy2(sorgente, accanto)
                    impostazioni = accanto
                except OSError:
                    impostazioni = None      # i dati sono salvi: basta e avanza
            accanto_cestino = os.path.splitext(destinazione)[0] + "_eliminati.json"
            try:
                if os.path.exists(cestino):
                    shutil.copy2(cestino, accanto_cestino)
                else:
                    with open(accanto_cestino, "w", encoding="utf-8") as fh:
                        json.dump([], fh)
                self.eliminati_nella_copia = True
            except OSError:
                pass                         # come sopra: il cestino e' un di piu'
        return destinazione, impostazioni, quanti

    def _estrai_copia_locale(self, percorso, cartella):
        """Tira fuori dati, impostazioni e cestino da una copia locale.

        Accetta sia lo zip - i tre file insieme - sia il vecchio .xlsx da solo,
        con i suoi _impostazioni.json e _eliminati.json accanto se ci sono: le
        copie salvate prima che lo zip esistesse devono restare ripristinabili.

        Ritorna (file dati, file impostazioni o None, file eliminati o None).
        """
        if not os.path.exists(percorso):
            raise InventoryError(T("Il file %s non esiste piu'.")
                                 % os.path.basename(percorso))
        if not percorso.lower().endswith(".zip"):
            radice = os.path.splitext(percorso)[0]
            accanto = radice + "_impostazioni.json"
            cestino = radice + "_eliminati.json"
            return (percorso,
                    accanto if os.path.exists(accanto) else None,
                    cestino if os.path.exists(cestino) else None)

        def per_nome(dentro, atteso):
            return next((n for n in dentro
                         if os.path.basename(n).lower() == atteso.lower()), None)

        try:
            with zipfile.ZipFile(percorso) as archivio:
                dentro = archivio.namelist()
                dati = next((n for n in dentro
                             if n.lower().endswith(".xlsx")
                             and not os.path.basename(n).startswith("~$")), None)
                if dati is None:
                    raise InventoryError(
                        T("Nella copia %s non c'e' nessun inventario.\n\n"
                          "Non e' stato ripristinato niente.")
                        % os.path.basename(percorso))
                eliminati = per_nome(dentro, NOME_ELIMINATI_NELLO_ZIP)
                impostazioni = per_nome(dentro, NOME_IMPOSTAZIONI_NELLO_ZIP)
                if impostazioni is None:
                    # copie vecchie: un solo .json, che erano le impostazioni
                    impostazioni = next((n for n in dentro
                                         if n.lower().endswith(".json")
                                         and n != eliminati), None)
                for dentro_lo_zip in (dati, impostazioni, eliminati):
                    if dentro_lo_zip:
                        archivio.extract(dentro_lo_zip, cartella)
        except zipfile.BadZipFile:
            raise InventoryError(
                T("%s non e' una copia leggibile: il file e' rovinato o non e'\n"
                  "uno di quelli salvati da questo programma.\n\n"
                  "Non e' stato ripristinato niente.") % os.path.basename(percorso))
        except OSError as exc:
            raise InventoryError(T("Non riesco a leggere la copia:\n%s") % exc)
        return (os.path.join(cartella, dati),
                os.path.join(cartella, impostazioni) if impostazioni else None,
                os.path.join(cartella, eliminati) if eliminati else None)

    def anteprima_copia_locale(self, percorso):
        """Che cosa c'e' dentro una copia locale, senza toccare niente.

        Serve al riepilogo: prima di riscrivere l'inventario di tutti bisogna
        poter leggere quanti dispositivi tornerebbero e con quali stanze.
        """
        from . import config

        cartella = tempfile.mkdtemp()
        try:
            dati, impostazioni, eliminati = self._estrai_copia_locale(percorso, cartella)
            try:
                dispositivi = InventoryStore(dati)._read()
            except InventoryError as exc:
                raise InventoryError(
                    T("%s non e' un inventario leggibile:\n%s\n\n"
                      "Non e' stato ripristinato niente.")
                    % (os.path.basename(percorso), exc))
            rapporto = {"dispositivi": len(dispositivi), "impostazioni": None,
                        "per_stanza": {}, "quando": None, "eliminati": 0}
            for it in dispositivi:
                stanza = clean(it.get("stanza")) or SENZA_STANZA
                rapporto["per_stanza"][stanza] = rapporto["per_stanza"].get(stanza, 0) + 1
            try:
                rapporto["quando"] = datetime.fromtimestamp(os.path.getmtime(dati))
            except OSError:
                pass
            if impostazioni:
                try:
                    with open(impostazioni, "r", encoding="utf-8") as fh:
                        rapporto["impostazioni"] = json.load(fh)
                except (OSError, ValueError):
                    rapporto["impostazioni"] = None
            if eliminati:
                try:
                    with open(eliminati, "r", encoding="utf-8") as fh:
                        voci = json.load(fh)
                    rapporto["eliminati"] = len(voci) if isinstance(voci, list) else 0
                except (OSError, ValueError):
                    rapporto["eliminati"] = 0
            return rapporto
        finally:
            shutil.rmtree(cartella, ignore_errors=True)

    def ripristina_da_copia_locale(self, percorso):
        """Rimette inventario e impostazioni come stanno in una copia locale.

        E' il ripristino per il caso peggiore: la cartella di rete sparita, con
        dentro i backup automatici. Da qui tornano anche le stanze, i tipi e le
        stanze con prestito, che nel solo file dei dispositivi non ci sono.

        Lo stato attuale viene salvato prima, cosi' anche questo si annulla.
        Ritorna (dispositivi, copia dello stato precedente, impostazioni si/no).
        """
        from . import config

        cartella = tempfile.mkdtemp()
        try:
            dati, impostazioni, eliminati = self._estrai_copia_locale(percorso, cartella)
            try:
                recuperati = InventoryStore(dati)._read()
            except InventoryError as exc:
                raise InventoryError(
                    T("%s non e' un inventario leggibile:\n%s\n\n"
                      "Non e' stato ripristinato niente.")
                    % (os.path.basename(percorso), exc))
            letta = None
            if impostazioni:
                try:
                    with open(impostazioni, "r", encoding="utf-8") as fh:
                        letta = json.load(fh)
                except (OSError, ValueError):
                    letta = None      # i dispositivi tornano lo stesso
            with _Lock(self.path):
                precedente = self.copia_di_sicurezza()
                try:
                    shutil.copy2(dati, self.path)
                except OSError as exc:
                    raise InventoryError(
                        T("Non riesco a ripristinare la copia:\n%s\n\n"
                          "L'inventario e' rimasto com'era.") % exc)
            # Le impostazioni vengono dopo i dati: se qui andasse storto
            # qualcosa, l'inventario e' comunque tornato al suo posto.
            if letta:
                try:
                    config.save_shared_config(self.path, letta)
                except OSError:
                    letta = None
            # E il cestino per ultimo, per la stessa ragione: se manca, quello
            # che c'e' adesso resta dov'e' invece di essere azzerato.
            self.eliminati_ripristinati = 0
            if eliminati:
                try:
                    with open(eliminati, "r", encoding="utf-8") as fh:
                        voci = json.load(fh)
                    if isinstance(voci, list):
                        self._scrivi_eliminati(self._pota_eliminati(voci))
                        self.eliminati_ripristinati = len(voci)
                except (OSError, ValueError):
                    pass
        finally:
            shutil.rmtree(cartella, ignore_errors=True)
        self.load()
        # Come per il ripristino da una copia automatica: chi e' tornato in
        # inventario esce dal cestino, o starebbe in tutti e due i posti.
        self.tolti_dal_ripristino = self.togli_dal_cestino()
        return len(recuperati), precedente, bool(letta)

    def copie_disponibili(self, quante=40):
        """Le copie di sicurezza, dalla piu' recente. (percorso, data, dispositivi)."""
        from . import config

        cartella = config.backup_dir()
        if not cartella or not os.path.isdir(cartella):
            return []
        trovate = []
        for nome in os.listdir(cartella):
            if not nome.lower().endswith(".xlsx") or nome.startswith("~$"):
                continue
            percorso = os.path.join(cartella, nome)
            try:
                quando = datetime.fromtimestamp(os.path.getmtime(percorso))
            except OSError:
                continue
            trovate.append((percorso, quando))
        trovate.sort(key=lambda voce: voce[1], reverse=True)
        elenco = []
        for percorso, quando in trovate[:quante]:
            try:
                quanti = len(InventoryStore(percorso)._read())
            except InventoryError:
                continue          # non e' un inventario leggibile: si salta
            elenco.append((percorso, quando, quanti))
        return elenco

    def restore(self, percorso):
        """Riporta l'inventario a una copia di sicurezza.

        Prima di sostituire il file viene salvata una copia dello stato attuale,
        cosi' anche un ripristino sbagliato si puo' annullare.

        Torna indietro anche il cestino salvato insieme a quella copia: sono lo
        stesso stato condiviso, e riportare i dispositivi senza gli eliminati
        lascerebbe nel cestino roba che nel frattempo e' rientrata in
        inventario. Delle copie vecchie, salvate prima che il cestino
        esistesse, quello che c'e' adesso resta com'e'.

        Ritorna (dispositivi ripristinati, copia dello stato precedente).
        `eliminati_ripristinati` dice quanti record sono tornati nel cestino.
        """
        if not os.path.exists(percorso):
            raise InventoryError(T("La copia %s non esiste piu'.") % os.path.basename(percorso))
        try:
            recuperati = InventoryStore(percorso)._read()
        except InventoryError as exc:
            raise InventoryError(
                T("%s non e' un inventario leggibile:\n%s\n\n"
                "Non e' stato ripristinato niente.") % (os.path.basename(percorso), exc))

        with _Lock(self.path):
            precedente = self.copia_di_sicurezza()
            try:
                shutil.copy2(percorso, self.path)
            except OSError as exc:
                raise InventoryError(
                    T("Non riesco a ripristinare la copia:\n%s\n\n"
                    "L'inventario e' rimasto com'era.") % exc)
        # Il cestino dopo i dati: se qui va storto qualcosa, l'inventario e'
        # comunque tornato al suo posto.
        self.eliminati_ripristinati = 0
        gemello = _cestino_accanto(percorso)
        if os.path.exists(gemello):
            try:
                with open(gemello, "r", encoding="utf-8") as fh:
                    voci = json.load(fh)
                if isinstance(voci, list):
                    self._scrivi_eliminati(self._pota_eliminati(voci))
                    self.eliminati_ripristinati = len(voci)
            except (OSError, ValueError):
                pass
        self.load()
        # Una copia vecchia puo' non avere il gemello, e allora il cestino di
        # adesso resta dov'e': dentro ci sarebbero dispositivi che il ripristino
        # ha appena rimesso in inventario, e da li' se ne "ripristinerebbe" una
        # seconda copia. Non possono stare in tutti e due i posti.
        self.tolti_dal_ripristino = self.togli_dal_cestino()
        return len(recuperati), precedente

    def reset(self):
        """Svuota l'inventario, tenendo solo cio' che non si potrebbe recuperare.

        Prima crea una copia di sicurezza del file dati. **Gli iPhone restano
        sempre**, in qualunque stato: non arrivano da un'importazione, quindi
        cancellarli qui significherebbe perderli per sempre. Restano dentro
        anche gli altri dispositivi non ancora eliminabili.

        Non parte se c'e' anche un solo **prestito aperto**: quel dispositivo
        e' in mano a qualcuno, e un reset lo toglierebbe dall'inventario senza
        nemmeno passare dal cestino. Prima si registrano i rientri.

        Ritorna (eliminati, mantenuti, percorso della copia).
        """

        def op(items):
            fuori = prestiti_aperti(items)
            if fuori:
                raise BloccoPrestitiAperti(
                    fuori, T("Non si puo' svuotare l'inventario"))
            copia = self.copia_di_sicurezza()
            tenuti = [it for it in items
                      if is_iphone(it.get("tipo")) or not puo_essere_eliminato(it)[0]]
            eliminati = len(items) - len(tenuti)
            items[:] = tenuti
            return eliminati, len(tenuti), copia

        return self._apply(op)

    def ship(self, tag):
        """Registra la spedizione al servizio telefonia."""
        tag = norm_tag(tag)

        def op(items):
            index = _index_of(items, tag)
            if index is None:
                raise InventoryError(T("Il dispositivo %s non esiste piu' nell'inventario.") % tag)
            item = items[index]
            if not is_iphone(item.get("tipo")):
                raise InventoryError(
                    T("La spedizione al servizio telefonia riguarda solo gli iPhone."))
            if is_shipped(item):
                raise InventoryError(
                    T("%s risulta gia' spedito il %s.") % (tag, item["spedito_il"]))
            item["spedito_il"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            normalize_state(item, self.stati)
            _stamp_item(item)
            return testo_spedizione(item)

        return self._apply(op)

    def rinomina_stanze(self, coppie):
        """Riscrive la stanza dei dispositivi che stavano in quelle rinominate.

        Senza questo, rinominare una stanza lasciava i dispositivi etichettati
        con un nome che non esiste piu': comparivano in una scheda a parte e
        bisognava spostarli a mano, uno per uno.

        Ritorna {nuovo nome: quanti dispositivi}, anche per le stanze che non
        ne avevano: chi legge il riepilogo vuole sapere che sono zero, non che
        non se ne e' parlato.
        """
        coppie = [(clean(v), clean(n)) for v, n in coppie if clean(v) and clean(n)]
        if not coppie:
            return {}

        self._rinomina_nel_cestino(coppie)

        def op(items):
            spostati = dict((nuovo, 0) for _v, nuovo in coppie)
            for vecchio, nuovo in coppie:
                for it in items:
                    if clean(it.get("stanza")) != vecchio:
                        continue
                    it["stanza"] = nuovo
                    _stamp_item(it)
                    spostati[nuovo] += 1
            return spostati

        return self._apply(op)

    def _rinomina_nel_cestino(self, coppie):
        """Rinominando una stanza, i record nel cestino la seguono.

        Senza questo, un dispositivo eliminato prima della rinomina tornerebbe
        in una stanza che non esiste piu': sparirebbe da tutte le schede. La
        stanza e' la stessa, ha solo cambiato nome, e il ripristino non deve
        chiedere niente a nessuno.
        """
        voci = self._leggi_eliminati()
        if not voci:
            return
        cambiate = False
        for voce in voci:
            for vecchio, nuovo in coppie:
                if clean(voce.get("stanza")) == vecchio:
                    voce["stanza"] = nuovo
                    cambiate = True
                scheda = voce.get("scheda") or {}
                if clean(scheda.get("stanza")) == vecchio:
                    scheda["stanza"] = nuovo
                    cambiate = True
        if cambiate:
            self._scrivi_eliminati(voci)

    def rinomina_tipi(self, coppie):
        """Riscrive il tipo dei dispositivi che avevano quello vecchio.

        Cambia solo il tipo: asset tag, numero di serie, stanza, stato, note,
        prestito e spedizione restano quelli che erano.

        Ritorna {nuovo nome: quanti dispositivi}.
        """
        coppie = [(clean(v), clean(n)) for v, n in coppie if clean(v) and clean(n)]
        if not coppie:
            return {}
        tocca = rinomina_tocca_gli_iphone(coppie)
        if tocca:
            raise InventoryError(
                T("Il tipo \"%s\" non si rinomina.\n\n"
                  "E' la parola con cui il programma riconosce i telefoni: da li'\n"
                  "vengono l'IMEI al posto dell'asset tag, la stanza obbligata, la\n"
                  "spedizione al servizio telefonia e il fatto che non si eliminino.\n"
                  "Cambiandola, i telefoni gia' registrati smetterebbero di essere\n"
                  "telefoni.") % tocca[0])

        def op(items):
            cambiati = dict((nuovo, 0) for _v, nuovo in coppie)
            for vecchio, nuovo in coppie:
                for it in items:
                    if clean(it.get("tipo")) != vecchio:
                        continue
                    it["tipo"] = nuovo
                    _stamp_item(it)
                    cambiati[nuovo] += 1
            return cambiati

        return self._apply(op)

    def move_to_room(self, tags, room):
        """Sposta i dispositivi selezionati.

        Ritorna (spostati, iphone_fermi, prestiti_fermi): gli iPhone stanno
        sempre nella loro stanza, e un dispositivo in prestito non si muove
        finche' non e' registrato il rientro.
        """
        wanted = set(norm_tag(t) for t in tags)
        room = clean(room)

        def op(items):
            moved = telefoni = prestati = 0
            for it in items:
                if it["asset_tag"] not in wanted:
                    continue
                if self.iphone_room and is_iphone(it.get("tipo")):
                    if it.get("stanza") != self.iphone_room:
                        it["stanza"] = self.iphone_room     # rimette a posto
                        _stamp_item(it)
                    telefoni += 1
                    continue
                if is_on_loan(it):
                    prestati += 1
                    continue
                if it.get("stanza") != room:
                    it["stanza"] = room
                    _stamp_item(it)
                    moved += 1
            return moved, telefoni, prestati

        return self._apply(op)

    def lend(self, tag, person):
        """Registra il prestito di un dispositivo a una persona."""
        tag = norm_tag(tag)
        person = clean(person)
        if not person:
            raise InventoryError(T("Indica il nome della persona a cui presti il dispositivo."))

        def op(items):
            index = _index_of(items, tag)
            if index is None:
                raise InventoryError(T("Il dispositivo %s non esiste piu' nell'inventario.") % tag)
            item = items[index]
            if is_iphone(item.get("tipo")):
                raise InventoryError(T("Gli iPhone non vengono dati in prestito."))
            if is_on_loan(item):
                raise InventoryError(
                    T("%s risulta gia' in prestito a %s dal %s.")
                    % (tag, item["prestato_a"], item["prestato_il"]))
            item["prestato_a"] = person
            item["prestato_il"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            normalize_state(item, self.stati)
            _stamp_item(item)
            return item["prestato_il"]

        return self._apply(op)

    def give_back(self, tag):
        """Chiude il prestito e rimette il dispositivo fra i disponibili."""
        tag = norm_tag(tag)

        def op(items):
            index = _index_of(items, tag)
            if index is None:
                raise InventoryError(T("Il dispositivo %s non esiste piu' nell'inventario.") % tag)
            item = items[index]
            if not is_on_loan(item):
                raise InventoryError(T("%s non risulta in prestito.") % tag)
            person = item["prestato_a"]
            item["prestato_a"] = ""
            item["prestato_il"] = ""
            item["stato"] = DISPONIBILE
            normalize_state(item, self.stati)
            _stamp_item(item)
            return person

        return self._apply(op)

    def set_stato(self, tag, stato):
        """Cambia lo stato di un dispositivo (tendina nell'elenco)."""
        tag = norm_tag(tag)
        stato = clean(stato)
        if stato not in self.stati:
            raise InventoryError(T("Stato non previsto: %s.") % stato)

        def op(items):
            index = _index_of(items, tag)
            if index is None:
                raise InventoryError(T("Il dispositivo %s non esiste piu' nell'inventario.") % tag)
            item = items[index]
            if is_iphone(item.get("tipo")):
                atteso = SPEDITO if is_shipped(item) else DA_RISPEDIRE
                raise InventoryError(
                    T("Lo stato degli iPhone e' sempre \"%s\" e non si cambia.") % atteso)
            if is_on_loan(item):
                raise InventoryError(
                    T("%s e' in prestito a %s: registra prima il rientro.")
                    % (tag, item["prestato_a"]))
            if item.get("stato") == stato:
                return False
            item["stato"] = stato
            _stamp_item(item)
            return True

        return self._apply(op)

    # Campi che si cambiano direttamente nell'elenco, senza aprire la scheda:
    # sono quelli che cambiano spesso e che si correggono guardando l'oggetto
    # che si ha in mano.
    CAMPI_AL_VOLO = ("note", "modello")

    def set_campo(self, tag, campo, valore):
        """Aggiorna un solo campo di testo (modifica al volo dall'elenco)."""
        if campo not in self.CAMPI_AL_VOLO:
            raise InventoryError(T("Il campo %s non si modifica dall'elenco.") % campo)
        tag = norm_tag(tag)
        valore = clean(valore)

        def op(items):
            index = _index_of(items, tag)
            if index is None:
                raise InventoryError(T("Il dispositivo %s non esiste piu' nell'inventario.") % tag)
            if items[index].get(campo, "") == valore:
                return False
            items[index][campo] = valore
            _stamp_item(items[index])
            return True

        return self._apply(op)

    def set_tipo(self, tag, tipo):
        """Cambia il tipo di un dispositivo (modifica al volo dall'elenco).

        Un iPhone non si trasforma in un laptop e viceversa: il passaggio
        cancellerebbe l'identificativo - l'IMEI da una parte, l'asset tag e il
        seriale dall'altra - e sarebbe una perdita silenziosa. Per quello serve
        eliminare e reinserire, che almeno si vede.
        """
        tag = norm_tag(tag)
        tipo = clean(tipo)

        def op(items):
            index = _index_of(items, tag)
            if index is None:
                raise InventoryError(T("Il dispositivo %s non esiste piu' nell'inventario.") % tag)
            attuale = items[index]
            if not tipo:
                raise InventoryError(T("Il tipo non puo' restare vuoto."))
            if clean(attuale.get("tipo")) == tipo:
                return False
            if is_iphone(attuale.get("tipo")) != is_iphone(tipo):
                raise InventoryError(
                    T("Non si passa da iPhone a un altro tipo, ne' viceversa.\n\n"
                    "Un iPhone e' identificato dall'IMEI, gli altri dispositivi "
                    "dall'asset tag\ne dal numero di serie: il passaggio "
                    "cancellerebbe l'identificativo.\n\n"
                    "Elimina il dispositivo e reinseriscilo con il tipo giusto."))
            attuale["tipo"] = tipo
            normalize_state(attuale, self.stati)
            _stamp_item(attuale)
            return True

        return self._apply(op)

    def set_note(self, tag, note):
        """Aggiorna soltanto le note (modifica al volo dall'elenco)."""
        return self.set_campo(tag, "note", note)

    def trova_duplicati(self):
        """Dispositivi registrati piu' di una volta con lo stesso identificativo.

        Il file dati e' un .xlsx che si puo' aprire e modificare a mano: e' da
        li' che i doppioni entrano, perche' il programma da solo non ne crea.
        Vengono segnalati anche i numeri di serie ripetuti, che duplicati non
        sono ma quasi sempre sono un errore di battitura.

        Ritorna (gruppi, seriali), dove gruppi e' una lista di (identificativo,
        [dispositivi]) e seriali una lista di (seriale, [dispositivi]).
        """
        items = self.load()
        per_tag, per_seriale = {}, {}
        for it in items:
            per_tag.setdefault(norm_tag(it.get("asset_tag")), []).append(it)
            seriale = clean(it.get("seriale")).upper()
            if seriale:
                per_seriale.setdefault(seriale, []).append(it)
        gruppi = [(tag, elenco) for tag, elenco in sorted(per_tag.items())
                  if tag and len(elenco) > 1]
        seriali = [(seriale, elenco) for seriale, elenco in sorted(per_seriale.items())
                   if len(elenco) > 1
                   and len(set(norm_tag(i.get("asset_tag")) for i in elenco)) > 1]
        return gruppi, seriali

    def _piu_recente(self, elenco):
        """Il doppione da tenere: l'ultimo toccato, o il primo se non si sa."""
        def quando(item):
            try:
                return datetime.strptime(item.get("modificato_il", ""),
                                         "%d/%m/%Y %H:%M:%S")
            except (ValueError, TypeError):
                return datetime.min
        return max(elenco, key=quando)

    def rimuovi_duplicati(self):
        """Tiene una sola copia di ogni dispositivo e toglie le altre.

        Si tiene la registrazione modificata piu' di recente: e' quella su cui
        qualcuno ha lavorato per ultimo. Gli iPhone protetti non si toccano
        nemmeno qui, e vengono elencati perche' se ne occupi una persona.

        Ritorna un rapporto con quello che e' stato fatto e quello che non si e'
        potuto fare.
        """
        gruppi, seriali = self.trova_duplicati()
        rapporto = {"gruppi": len(gruppi), "eliminati": [], "tenuti": [],
                    "protetti": [], "seriali": seriali, "copia": None,
                    "prima": len(self.items), "dopo": len(self.items)}
        if not gruppi:
            return rapporto

        da_togliere = []
        for tag, elenco in gruppi:
            tenuto = self._piu_recente(elenco)
            rapporto["tenuti"].append(tenuto)
            for item in elenco:
                if item is tenuto:
                    continue
                libero, sblocco = puo_essere_eliminato(item)
                if not libero:
                    motivo = ("iPhone non ancora rispedito" if sblocco is None
                              else "in conservazione fino al %s"
                              % sblocco.strftime("%d/%m/%Y"))
                    rapporto["protetti"].append((item, motivo))
                    continue
                da_togliere.append(item)
        if not da_togliere:
            return rapporto

        rapporto["copia"] = self.copia_di_sicurezza()

        # I doppioni si riconoscono per contenuto, non per identita': due copie
        # identiche restano due righe distinte nel file, e se ne toglie una sola
        # per ogni copia in eccesso.
        chiavi = [(norm_tag(i.get("asset_tag")), i.get("modificato_il"),
                   i.get("modificato_da")) for i in da_togliere]

        def op(items):
            rimasti, tolti = [], []
            residui = list(chiavi)
            for it in items:
                chiave = (norm_tag(it.get("asset_tag")), it.get("modificato_il"),
                          it.get("modificato_da"))
                if chiave in residui:
                    residui.remove(chiave)
                    tolti.append(it)
                    continue
                rimasti.append(it)
            items[:] = rimasti
            return tolti

        tolti = self._apply(op)
        rapporto["eliminati"] = tolti
        rapporto["dopo"] = len(self.items)
        rapporto["prima"] = rapporto["dopo"] + len(tolti)
        return rapporto

    def anteprima_eliminazione(self, codici):
        """Che cosa succederebbe eliminando i codici indicati, senza toccarli.

        Serve all'eliminazione in blocco: prima di cancellare trenta righe
        bisogna poter leggere che cosa sparisce e da dove. Un codice puo'
        arrivare incollato da Excel come riga intera, quindi si prova ogni
        pezzo della riga: il primo che corrisponde a un dispositivo vince.

        Ritorna (da_eliminare, non_trovati, bloccati), dove bloccati e' una
        lista di (dispositivo, motivo).
        """
        items = self.load()
        per_tag = {}
        for it in items:
            per_tag[norm_tag(it["asset_tag"])] = it
        da_eliminare, non_trovati, bloccati = [], [], []
        gia_visti = set()
        for riga in codici:
            trovato = None
            for pezzo in _pezzi_di_riga(riga):
                trovato = per_tag.get(norm_tag(pezzo))
                if trovato is not None:
                    break
            if trovato is None:
                if clean(riga):
                    non_trovati.append(clean(riga))
                continue
            tag = norm_tag(trovato["asset_tag"])
            if tag in gia_visti:
                continue                  # la stessa riga incollata due volte
            gia_visti.add(tag)
            if is_on_loan(trovato):
                bloccati.append((trovato, T("in prestito a %s: registra prima il rientro")
                                 % trovato.get("prestato_a", "")))
                continue
            libero, sblocco = puo_essere_eliminato(trovato)
            if not libero:
                if sblocco is None:
                    motivo = T("iPhone non ancora rispedito al servizio telefonia")
                else:
                    motivo = T("in conservazione fino al %s") % sblocco.strftime("%d/%m/%Y")
                bloccati.append((trovato, motivo))
                continue
            da_eliminare.append(trovato)
        return da_eliminare, non_trovati, bloccati

    def anteprima_importazione(self, incoming, mode="merge", stanza=None):
        """Che cosa succederebbe importando, senza scrivere niente.

        Un riepilogo con dei numeri soltanto - "12 aggiunti, 3 saltati" - non
        dice dove finiranno i dispositivi ne' che cosa si sta per perdere. Qui si
        contano le aggiunte e i salti stanza per stanza, e si dice quanti
        dispositivi ci saranno alla fine.

        Un asset tag gia' in inventario non viene importato: la riga si salta e
        si dice dov'e' quello che c'e' gia'. Chi sta importando decide se e' un
        errore di battitura o il dispositivo che aveva gia' registrato, ma
        l'importazione non gli riscrive la scheda sotto senza dirglielo.
        """
        items = self.load()
        # In sostituzione il confronto si fa con quello che sopravvive alla
        # pulizia, non con l'inventario di adesso: il resto sta per sparire.
        sopravvissuti = items
        if mode == "replace":
            sopravvissuti = [it for it in items
                             if is_iphone(it.get("tipo"))
                             or (stanza is not None and it.get("stanza") != stanza)]
        dove_sta = dict((norm_tag(i["asset_tag"]), i) for i in sopravvissuti)
        presenti = set(dove_sta)
        per_stanza = {}
        senza_identificativo = 0
        senza_stanza = 0
        gia_presenti = []
        nuovi = set()
        for raw in incoming:
            item = dict(raw)
            tag = norm_tag(item.get("asset_tag"))
            if not tag:
                senza_identificativo += 1
                continue
            dove = clean(stanza) if stanza is not None else clean(item.get("stanza"))
            if not dove and self.iphone_room and is_iphone(item.get("tipo")):
                dove = clean(self.iphone_room)
            # il nome ufficiale, come quello con cui entrera' davvero: l'anteprima
            # deve mostrare la stanza in cui il dispositivo si trovera', non
            # quella scritta a modo suo da chi ha preparato il foglio
            dove = self.stanza_canonica(dove)
            if not dove:
                # senza stanza non si importa: si dice prima, non dopo
                senza_stanza += 1
                continue
            riga = per_stanza.setdefault(dove, {"aggiunti": 0, "saltati": 0})
            if tag in presenti:
                riga["saltati"] += 1
                trovato = dove_sta.get(tag)
                gia_presenti.append(
                    {"asset_tag": tag,
                     "stanza": clean((trovato or {}).get("stanza")) or SENZA_STANZA})
            elif tag not in nuovi:
                # ripetuto nel foglio: entrera' una volta sola, vale l'ultima riga
                riga["aggiunti"] += 1
                nuovi.add(tag)

        fuori = prestiti_aperti(items, stanza) if mode == "replace" else []
        eliminati = 0
        if mode == "replace":
            eliminati = sum(1 for it in items
                            if not is_iphone(it.get("tipo"))
                            and (stanza is None or it.get("stanza") == stanza)
                            and puo_essere_eliminato(it)[0])
        aggiunti = sum(r["aggiunti"] for r in per_stanza.values())
        return {
            "per_stanza": per_stanza,
            "aggiunti": aggiunti,
            "saltati": sum(r["saltati"] for r in per_stanza.values()),
            "gia_presenti": gia_presenti,
            "eliminati": eliminati,
            "senza_identificativo": senza_identificativo,
            "senza_stanza": senza_stanza,
            "prestiti_aperti": fuori,
            "prima": len(items),
            "dopo": len(items) - eliminati + aggiunti,
        }

    def import_items(self, incoming, mode="merge", stanza=None):
        """Carica i dispositivi letti da un file.

        mode: 'merge' aggiunge e aggiorna, 'replace' svuota prima di caricare.
        stanza: se indicata, l'importazione riguarda solo quella stanza - tutte
            le righe le vengono assegnate, e una sostituzione tocca solo lei.

        Prima di una sostituzione viene salvata una copia del file dati. Gli
        iPhone non si toccano mai: non arrivano da un'importazione, quindi una
        sostituzione li cancellerebbe senza possibilita' di recupero.

        Una sostituzione non parte se fra i dispositivi che verrebbero tolti
        c'e' un **prestito aperto**: quel dispositivo e' in mano a qualcuno.
        Prima si registrano i rientri.

        Un asset tag gia' in inventario **non viene importato**: la riga si
        salta e finisce in `gia_presenti`, con la stanza in cui sta quello che
        c'e' gia'. Un'importazione non riscrive di nascosto la scheda di un
        dispositivo gia' registrato: se e' lo stesso non serve, e se e' un altro
        e' un errore di battitura da guardare.

        Chi era solo nel cestino invece entra: quello non e' in inventario. La
        sua voce fra gli eliminati di recente viene tolta subito dopo, o il
        dispositivo resterebbe in tutti e due i posti.

        Un dispositivo **senza stanza non entra**: non si saprebbe dove
        cercarlo, non comparirebbe in nessuna stanza e nessuna esportazione per
        stanza lo conterrebbe. La riga si salta e finisce in `senza_stanza`. Chi
        importa dall'interfaccia la stanza se l'e' gia' vista chiedere; questo
        e' il muro che regge anche se qualcuno chiama l'archivio direttamente.

        Ritorna un dizionario con aggiunti, gia_presenti, senza_stanza,
        eliminati, copia e tolti_dal_cestino.
        """

        def op(items):
            esito = {"aggiunti": 0, "gia_presenti": [], "senza_stanza": [],
                     "eliminati": 0, "copia": None}
            if mode == "replace":
                fuori = prestiti_aperti(items, stanza)
                if fuori:
                    raise BloccoPrestitiAperti(
                        fuori,
                        T("Non si puo' sostituire %s")
                        % (stanza or T("tutto l'inventario")))
                esito["copia"] = self.copia_di_sicurezza()
                prima = len(items)
                items[:] = [it for it in items
                            if is_iphone(it.get("tipo"))
                            or (stanza is not None and it.get("stanza") != stanza)]
                esito["eliminati"] = prima - len(items)
            # Chi c'era gia' prima di questa importazione: e' con loro che si
            # fa il confronto. Le righe aggiunte adesso non entrano nel
            # paragone, o un foglio che ripete lo stesso identificativo si
            # salterebbe da solo la seconda volta - e li' vale un'altra regola,
            # quella del foglio: l'ultima riga vince.
            c_erano_prima = {it["asset_tag"]: it for it in items}
            appena_messi = {}
            for raw in incoming:
                item = dict(raw)
                item["asset_tag"] = norm_tag(item.get("asset_tag"))
                if stanza is not None:
                    item["stanza"] = clean(stanza)
                if not item["asset_tag"]:
                    continue
                normalize_identity(item)
                gia = c_erano_prima.get(item["asset_tag"])
                if gia is not None:
                    esito["gia_presenti"].append(
                        {"asset_tag": item["asset_tag"],
                         "stanza": clean(gia.get("stanza")) or SENZA_STANZA})
                    continue
                self._enforce_iphone_room(item)
                # la stanza entra sempre col nome ufficiale: scritta a modo suo
                # da chi ha preparato il foglio, sarebbe una stanza in piu'
                item["stanza"] = self.stanza_canonica(item.get("stanza"))
                if not item["stanza"]:
                    esito["senza_stanza"].append(item["asset_tag"])
                    continue
                normalize_state(item, self.stati)
                _stamp_item(item)
                if item["asset_tag"] in appena_messi:
                    items[appena_messi[item["asset_tag"]]] = item
                else:
                    appena_messi[item["asset_tag"]] = len(items)
                    items.append(item)
                    esito["aggiunti"] += 1
            return esito

        esito = self._apply(op)
        # Il cestino si ripulisce dopo, sui soli identificativi appena entrati:
        # erano eliminati, adesso sono di nuovo in inventario.
        esito["tolti_dal_cestino"] = self.togli_dal_cestino()
        return esito


# ------------------------------------------------------------- utilita'


SENZA_STANZA = "(senza stanza)"


def _pezzi_di_riga(riga):
    """I possibili identificativi in una riga incollata da Excel.

    Incollando da un foglio si porta dietro tutta la riga, separata da
    tabulazioni; incollando una colonna sola arriva un codice per riga. Vanno
    bene tutti e due.
    """
    testo = str(riga or "")
    for separatore in ("\t", ";", ","):
        testo = testo.replace(separatore, "\x00")
    return [p.strip() for p in testo.split("\x00") if p.strip()]


def _quanti_cambiati(prima, dopo):
    """Quanti record sono stati aggiunti, tolti o modificati da un'operazione.

    Si contano i record, non le operazioni: eliminarne trenta in un colpo solo
    e' un cambiamento grosso quanto trenta eliminazioni una per una, e chi
    tiene una copia locale vuole saperlo.
    """
    adesso = dict((it["asset_tag"], it) for it in dopo)
    cambiati = len(set(prima) ^ set(adesso))
    for tag in set(prima) & set(adesso):
        if prima[tag] != adesso[tag]:
            cambiati += 1
    return cambiati


def _index_of(items, tag):
    for i, it in enumerate(items):
        if it["asset_tag"] == tag:
            return i
    return None


def _stamp_item(item):
    # con i secondi due inserimenti nello stesso minuto restano in ordine
    item["modificato_il"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    item["modificato_da"] = current_user()


RIGHE_DA_ISPEZIONARE = 12


def _trova_intestazioni(rows):
    """La riga delle intestazioni, saltando eventuali titoli in cima al foglio.

    I file esportati dal programma per una singola stanza hanno il nome della
    stanza e la data prima della tabella; lo stesso capita spesso ai fogli che
    arrivano da altri sistemi.
    """
    for numero, row in enumerate(rows):
        if numero >= RIGHE_DA_ISPEZIONARE:
            break
        if row is None or all(c is None or clean(c) == "" for c in row):
            continue
        mapping = map_headers(row)
        if {"asset_tag", "imei"} & set(mapping.values()):
            return row, mapping
    return None, {}


def map_headers(header_row):
    """Mappa indice colonna -> campo, riconoscendo le intestazioni note."""
    mapping = {}
    for idx, cell in enumerate(header_row or ()):
        name = clean(cell).lower()
        if not name:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if field in mapping.values():
                continue
            if name in aliases:
                mapping[idx] = field
                break
    return mapping


def _row_to_item(row, mapping):
    item = new_item()
    for idx, field in mapping.items():
        if idx < len(row):
            item[field] = norm_tag(row[idx]) if field == "asset_tag" else clean(row[idx])
    normalize_identity(item)
    normalize_state(item)
    return item if item["asset_tag"] else None


# Un file da cui si eliminano dispositivi non ha bisogno di essere un
# inventario: se ne leggono al massimo queste righe, oltre non e' piu' un
# elenco preparato a mano ma un file sbagliato.
RIGHE_MASSIME_ELIMINAZIONE = 5000


def righe_da_workbook(path):
    """Le righe di un file Excel, come testo, per l'eliminazione in blocco.

    Va bene un'esportazione del programma - se trova le intestazioni prende
    solo le colonne Asset Tag e IMEI - ma va bene anche una colonna di codici
    incollata a mano e salvata senza intestazioni: in quel caso vale tutta la
    riga, e a riconoscere il codice ci pensa l'anteprima, che prova ogni pezzo.
    Chi prepara un elenco di cose da eliminare non deve doverlo formattare.

    Ritorna (righe, esito), con esito = {"fogli", "vuote", "troncato"}.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InventoryError(T("Impossibile leggere il file:\n%s") % exc)
    try:
        righe = []
        esito = {"fogli": 0, "vuote": 0, "troncato": False}
        for ws in wb.worksheets:
            tutte = list(ws.iter_rows(values_only=True))
            if not tutte:
                continue
            esito["fogli"] += 1
            header, mapping = _trova_intestazioni(iter(tutte))
            colonne = [i for i, campo in mapping.items()
                       if campo in ("asset_tag", "imei")] if header else []
            saltare = True if header is not None else False
            for row in tutte:
                if row is None or all(c is None or clean(c) == "" for c in row):
                    esito["vuote"] += 1
                    continue
                if saltare:
                    # la riga delle intestazioni non e' un dispositivo
                    if row is header or list(row) == list(header):
                        saltare = False
                        continue
                    continue          # i titoli sopra la tabella
                if colonne:
                    pezzi = [clean(row[i]) for i in colonne if i < len(row)]
                else:
                    pezzi = [clean(c) for c in row]
                pezzi = [p for p in pezzi if p]
                if not pezzi:
                    esito["vuote"] += 1
                    continue
                righe.append("\t".join(pezzi))
                if len(righe) >= RIGHE_MASSIME_ELIMINAZIONE:
                    esito["troncato"] = True
                    return righe, esito
        return righe, esito
    finally:
        wb.close()


def rows_from_workbook(path, rooms=None):
    """Legge un file .xlsx/.xlsm esterno.

    Ritorna (items, esito), dove esito e' un dizionario con le righe scartate,
    quelle che hanno preso la stanza da un separatore, gli iPhone ignorati, le
    colonne non riconosciute, le righe senza modello e l'elenco delle stanze
    per cui e' comparso un separatore.

    Se nel foglio compaiono righe-separatore con il nome (o l'abbreviazione) di
    una stanza, tutte le righe successive fino al separatore seguente vengono
    assegnate a quella stanza: e' il modo per dividere per stanza un unico
    inventario. Gli iPhone non si importano: sono gestiti solo a mano.
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InventoryError(T("Impossibile leggere il file:\n%s") % exc)
    try:
        tags = tag_stanze(rooms or [])
        items = []
        esito = {"scartate": 0, "da_tag": 0, "iphone": 0, "senza_modello": 0,
                 "stanze_trovate": [], "colonne_ignorate": [], "doppioni": []}
        visti = set()
        letto = False

        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header, mapping = _trova_intestazioni(rows)
            if header is None:
                continue          # un foglio di istruzioni o di appunti: si salta
            letto = True
            for cella in header:
                nome = clean(cella)
                if nome and nome not in esito["colonne_ignorate"] \
                        and list(header).index(cella) not in mapping:
                    esito["colonne_ignorate"].append(nome)

            # Un foglio intitolato come una stanza vale come separatore: e' la
            # forma che prende un'esportazione con un foglio per stanza.
            stanza_corrente = tags.get(clean(ws.title).upper())
            if stanza_corrente and stanza_corrente not in esito["stanze_trovate"]:
                esito["stanze_trovate"].append(stanza_corrente)

            for row in rows:
                if row is None or all(c is None or clean(c) == "" for c in row):
                    continue
                stanza = riga_tag(row, tags)
                if stanza is None:
                    stanza = separatore_con_avanzi(row, mapping, tags)
                if stanza is not None:
                    stanza_corrente = stanza
                    if stanza not in esito["stanze_trovate"]:
                        esito["stanze_trovate"].append(stanza)
                    continue
                item = _row_to_item(row, mapping)
                if not item:
                    esito["scartate"] += 1
                    continue
                if is_iphone(item.get("tipo")):
                    esito["iphone"] += 1
                    continue
                if stanza_corrente:
                    item["stanza"] = stanza_corrente
                    esito["da_tag"] += 1
                if not item.get("modello"):
                    esito["senza_modello"] += 1
                if item["asset_tag"] in visti:
                    # due righe con lo stesso identificativo: vale l'ultima, ma
                    # chi importa deve sapere che il foglio ne conteneva due
                    esito["doppioni"].append(item["asset_tag"])
                visti.add(item["asset_tag"])
                items.append(item)

        if not letto:
            raise InventoryError(
                T("Nel file non e' stata trovata la colonna \"Asset Tag\" (o \"IMEI\").\n"
                "Ci deve essere una riga con le intestazioni delle colonne.")
            )
        return items, esito
    finally:
        wb.close()


# Excel misura le colonne in caratteri del font predefinito, non in pixel. La
# corrispondenza non e' esatta - le lettere non sono tutte larghe uguali - per
# questo si aggiunge un margine invece di fidarsi del conteggio secco.
MARGINE_COLONNA = 3
LARGHEZZA_COLONNA_MIN = 9
LARGHEZZA_COLONNA_MAX = 80


def larghezza_colonna(titolo, valori):
    """Quanto deve essere larga una colonna del foglio perche' si legga tutta.

    Le larghezze fisse tagliavano il testo, o lo facevano sbordare sulla cella
    accanto: chi apre il file si trova a doverle allargare a mano una per una.
    Si misura invece il contenuto vero, e non si scende mai sotto il titolo -
    una colonna vuota deve comunque dire che cosa conterrebbe.
    """
    piu_lungo = max([len(str(v)) for v in valori if v] or [0])
    largo = max(len(titolo), piu_lungo) + MARGINE_COLONNA
    return min(max(largo, LARGHEZZA_COLONNA_MIN), LARGHEZZA_COLONNA_MAX)


def _style_sheet(ws, row_count):
    """Formattazione minima del file dati (l'export di stampa e' piu' curato)."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    # anche il file dati si apre in Excel: le colonne si misurano sul contenuto
    for i, field in enumerate(ALL_FIELDS, start=1):
        valori = [ws.cell(row=r, column=i).value for r in range(2, row_count + 2)]
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            larghezza_colonna(HEADERS[field], valori)
    if row_count:
        ws.auto_filter.ref = "A1:%s%d" % (
            ws.cell(row=1, column=len(ALL_FIELDS)).column_letter, row_count + 1)
