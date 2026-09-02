"""Esportazione e stampa in formato Excel."""

import os
import subprocess
import sys
import zipfile
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .store import (ALL_FIELDS, HEADERS, InventoryError, NON_DISPONIBILE,
                    SPEDITO, STATI, clean, is_iphone, larghezza_colonna,
                    valore_visibile)
from .lingua import T, intestazione, stato as traduci_stato

PRINT_FIELDS = ["asset_tag", "tipo", "modello", "seriale", "imei", "restituito_da",
                "stanza", "stato", "prestato_a", "prestato_il", "spedito_il", "note"]
PRINT_WIDTHS = {"asset_tag": 16, "tipo": 10, "modello": 26, "seriale": 16,
                "imei": 18, "restituito_da": 20, "stanza": 22, "stato": 24,
                "prestato_a": 20, "prestato_il": 15, "spedito_il": 15, "note": 24,
                "modificato_il": 18, "modificato_da": 24}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_BAND_FILL = PatternFill("solid", fgColor="F2F6FA")
_LOAN_FILL = PatternFill("solid", fgColor="FDEEEC")
_LOAN_TEXT = "A93226"
_SHIP_FILL = PatternFill("solid", fgColor="F6EFFB")
_SHIP_TEXT = "6C3483"
_THIN = Side(style="thin", color="B7C4D2")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _safe_sheet_title(name):
    for ch in "[]:*?/\\":
        name = name.replace(ch, "-")
    return (name or "Foglio")[:31]


def _write_table(ws, items, fields, title=None, subtitle=None, lingua=None):
    row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
        row = 2
        if subtitle:
            cell = ws.cell(row=2, column=1, value=subtitle)
            cell.font = Font(size=9, color="666666")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(fields))
            row = 3
        row += 1  # riga vuota di respiro

    header_row = row
    for col, field in enumerate(fields, start=1):
        cell = ws.cell(row=header_row, column=col,
                       value=intestazione(HEADERS[field], lingua))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = _BORDER
    ws.row_dimensions[header_row].height = 20

    for i, item in enumerate(items):
        r = header_row + 1 + i
        on_loan = item.get("stato") == NON_DISPONIBILE
        spedito = item.get("stato") == SPEDITO
        for col, field in enumerate(fields, start=1):
            valore = valore_visibile(item, field)
            if field == "stato":
                valore = traduci_stato(valore, lingua)
            cell = ws.cell(row=r, column=col, value=valore)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "note"))
            if on_loan:
                cell.fill = _LOAN_FILL
                cell.font = Font(color=_LOAN_TEXT)
            elif spedito:
                cell.fill = _SHIP_FILL
                cell.font = Font(color=_SHIP_TEXT)
            elif i % 2:
                cell.fill = _BAND_FILL

    for col, field in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(col)].width = larghezza_colonna(
            intestazione(HEADERS[field], lingua),
            [traduci_stato(valore_visibile(i, field), lingua) if field == "stato"
             else valore_visibile(i, field) for i in items])

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if items:
        last = get_column_letter(len(fields))
        ws.auto_filter.ref = "A%d:%s%d" % (header_row, last, header_row + len(items))
    return header_row


def _setup_print(ws, header_row, fields, footer_left):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "%d:%d" % (header_row, header_row)
    ws.print_area = "A1:%s%d" % (get_column_letter(len(fields)), ws.max_row)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    ws.oddFooter.left.text = footer_left
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.text = "Pagina &P di &N"
    ws.oddFooter.right.size = 8


def _sottotitolo(stamp, quanti, lingua):
    if lingua == "en":
        return "Exported on %s - %d devices" % (stamp, quanti)
    return "Esportato il %s - %d dispositivi" % (stamp, quanti)


# Le colonne portanti di un inventario: ci sono sempre, anche vuote, perche'
# danno al file la forma che chi lo apre - o chi lo reimporta - si aspetta.
CAMPI_PORTANTI = ["asset_tag", "tipo", "modello", "seriale", "stanza", "stato", "note"]

# Un file esportato dice che cosa abbiamo, dove sta e che cosa c'e' da sapere:
# le note viaggiano con il dispositivo, perche' sono quello che una riga ha di
# particolare. Il resto - lo stato, il modello, il numero di serie, i prestiti,
# l'IMEI, chi ha toccato la riga per ultimo - serve a chi lavora davanti
# all'elenco, dentro la stanza che lo riguarda, e non a chi riceve il file.
#
# Attenzione: da un'esportazione non si ricostruisce un inventario, perche' quei
# campi non ci sono. Per quello c'e' la copia locale, che copia il file vero.
CAMPI_ESPORTAZIONE = ["asset_tag", "tipo", "stanza", "note"]


def campi_con_valore(items, fields):
    """Toglie le colonne che in questo file resterebbero vuote in ogni riga.

    Un'esportazione non contiene iPhone, quindi non ha nessun IMEI da scrivere;
    l'esportazione di una stanza senza prestiti non ha prestiti. Colonne di sole
    caselle vuote allargano il foglio e non dicono niente a chi lo legge.

    Le colonne portanti restano comunque: un inventario deve avere sempre lo
    stesso impianto, altrimenti due file dello stesso inventario non si
    somigliano.
    """
    if not items:
        return [f for f in fields if f in CAMPI_PORTANTI]
    return [f for f in fields
            if f in CAMPI_PORTANTI or any(clean(i.get(f)) for i in items)]


def _sort_key(item):
    return (item.get("stanza", ""), item.get("asset_tag", ""))


def export(items, path, group_by_room=False, rooms=None, for_print=False,
           con_iphone=False, titolo=None, lingua=None):
    """Scrive l'inventario in un file .xlsx.

    group_by_room: un foglio per stanza (piu' un foglio con il totale).
    for_print: aggiunge titolo, intestazioni ripetute e impaginazione A4, e
        porta anche i dati di consultazione - prestiti, IMEI, spedizioni -
        perche' la stampa serve a chi lavora, non a chi riceve il file.
    con_iphone: gli iPhone sono gestiti solo a mano e restano fuori dalle
        esportazioni; la stampa interna invece li include.
    titolo: se indicato, compare in testa al foglio e ne diventa il nome. Serve
        all'esportazione di una singola stanza, che deve dichiararsi.
    lingua: "en" per scrivere intestazioni e stati in inglese. I nomi delle
        stanze e dei tipi restano come li ha scritti l'utente.
    """
    fields = list(PRINT_FIELDS) if for_print else list(CAMPI_ESPORTAZIONE)
    if not con_iphone:
        items = [i for i in items if not is_iphone(i.get("tipo"))]
    items = sorted(items, key=_sort_key)
    # si decide una volta sola, su tutto quello che finira' nel file: i fogli di
    # uno stesso documento devono avere le stesse colonne
    fields = campi_con_valore(items, fields)
    stamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name, subset, room=None):
        # La colonna Stanza resta sempre: un foglio deve dire da solo di chi e',
        # anche se viene copiato o stampato fuori dal suo file.
        sheet_fields = list(fields)
        ws = wb.create_sheet(_safe_sheet_title(name))
        title = subtitle = None
        if for_print:
            title = 'Site Services : Inventario Iphone, Laptop e Tablet' + (" - %s" % room if room else "")
            subtitle = "Stampato il %s - %d dispositivi" % (stamp, len(subset))
        elif room:
            title = room
            subtitle = _sottotitolo(stamp, len(subset), lingua)
        elif titolo:
            title = titolo
            subtitle = _sottotitolo(stamp, len(subset), lingua)
        header_row = _write_table(ws, subset, sheet_fields, title, subtitle, lingua)
        if for_print:
            _setup_print(ws, header_row, sheet_fields,
                         'Site Services : Inventario Iphone, Laptop e Tablet' + " - %s" % stamp)

    if group_by_room:
        room_names = list(rooms or [])
        for room in sorted(set(room_names) | set(i.get("stanza", "") for i in items)):
            subset = [i for i in items if i.get("stanza", "") == room]
            add_sheet(room or "Senza stanza", subset, room=room or "Senza stanza")
        if not wb.sheetnames:
            add_sheet("Inventario", items)
    else:
        add_sheet(titolo or "Inventario", items)

    try:
        wb.save(path)
    except Exception as exc:
        raise InventoryError("Impossibile scrivere %s:\n%s" % (path, exc))
    finally:
        wb.close()
    return path


def export_per_stanza(items, cartella, rooms=None, con_iphone=False, lingua=None):
    """Un file separato per ogni stanza che abbia dispositivi.

    Ritorna l'elenco dei percorsi scritti.
    """
    if not con_iphone:
        items = [i for i in items if not is_iphone(i.get("tipo"))]
    stanze = list(rooms or [])
    for item in items:
        if item.get("stanza") and item["stanza"] not in stanze:
            stanze.append(item["stanza"])
    giorno = datetime.now().strftime("%Y%m%d")
    scritti = []
    for stanza in stanze:
        subset = [i for i in items if i.get("stanza", "") == stanza]
        if not subset:
            continue
        nome = "Inventario_%s_%s.xlsx" % (
            "_".join("".join(c if c.isalnum() or c in " -_" else "-"
                             for c in str(stanza)).split()) or "stanza", giorno)
        scritti.append(export(subset, os.path.join(cartella, nome),
                              rooms=[stanza], titolo=stanza, con_iphone=con_iphone,
                              lingua=lingua))
    return scritti


def build_print_file(items, group_by_room=False, rooms=None):
    """Genera in una cartella temporanea il file pronto da stampare."""
    name = "Inventario_stampa_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(tempfile.gettempdir(), name)
    return export(items, path, group_by_room=group_by_room, rooms=rooms,
                  for_print=True, con_iphone=True)


# Il modello ha esattamente le colonne di un file esportato: si esporta, si
# corregge in Excel, si reimporta. Due formati diversi per la stessa cosa
# obbligano a ricordarsi quale sia quale, e prima o poi si sbaglia.
TEMPLATE_FIELDS = list(CAMPI_ESPORTAZIONE)

# Un valore d'esempio per ogni colonna, che serve solo a darle una larghezza
# sensata: il modello e' vuoto, quindi non c'e' contenuto da misurare, ma chi
# scrive dentro deve poterci stare comodo.
TEMPLATE_ESEMPI = {
    "asset_tag": "IT-0000",
    "tipo": "Laptop",
    "stanza": "Magazzino Disaster Recovery",
    "note": "Batteria da sostituire, rientro dal reparto",
    "stato": "Guasto in attesa tecnico",
    "modello": "Lenovo ThinkPad T14 Gen 5",
    "seriale": "PF4A1B2C",
}
TEMPLATE_TIPI = ["Laptop", "Tablet"]
RIGHE_PER_STANZA = 8


ISTRUZIONI_IT = [
    ("Modello di inventario - laptop e tablet", True),
    ("", False),
    ('Compila il foglio "Inventario" e importalo dal programma con', False),
    ("Importa xls...  Le righe azzurre con il nome della stanza sono", False),
    ("separatori: tutto cio' che scrivi sotto una di esse finisce in", False),
    ("quella stanza, fino al separatore successivo.", False),
    ("", False),
    ("Regole", True),
    ("- L'Asset Tag e' l'unico dato indispensabile: senza, la riga viene", False),
    ("  scartata. Tipo, stanza e note si possono lasciare vuoti.", False),
    ("- Il modello ha le stesse colonne di un file esportato: puoi", False),
    ("  esportare, correggere in Excel e reimportare.", False),
    ("- Servono altre colonne? Aggiungile: modello, numero di serie e", False),
    ("  stato vengono riconosciuti dal nome e importati lo stesso.", False),
    ("- L'asset tag identifica il dispositivo: importando due volte lo", False),
    ("  stesso asset tag, la scheda viene aggiornata invece che duplicata.", False),
    ("- Tipo e Stanza hanno la tendina: usa i valori proposti.", False),
    ("- La stanza si puo' anche non scriverla: la dicono i separatori.", False),
    ("", False),
    ("Copia e incolla", True),
    ("- Da qui puoi copiare intere colonne e incollarle nel programma:", False),
    ("  in Elimina + si incollano gli asset tag, uno per riga.", False),
    ("- Nel programma tutti i campi accettano Ctrl+C e Ctrl+V, e il", False),
    ("  tasto destro apre Copia / Incolla.", False),
    ("- Dall'elenco del programma si copia la riga selezionata con", False),
    ("  Ctrl+C, o l'identificativo da solo con il tasto destro: si", False),
    ("  incolla qui dentro cosi' com'e'.", False),
    ("- Puoi aggiungere righe sotto un separatore, o spostare i separatori.", False),
    ("- Non cambiare i nomi delle colonne nella prima riga.", False),
    ("- Le righe lasciate vuote vengono semplicemente ignorate.", False),
    ("", False),
    ("Valori ammessi", True),
    ("TIPI", False),
    ("STATI", False),
    ("Le tendine funzionano in Excel e in LibreOffice. Numbers non le", False),
    ("importa: in quel caso scrivi i valori qui sopra, sono gli stessi.", False),
    ("", False),
    ("Gli iPhone non si importano", True),
    ("Sono gestiti solo a mano dal programma e non compaiono ne' nelle", False),
    ("importazioni ne' nelle esportazioni. Se ne inserisci qui, vengono", False),
    ("ignorati.", False),
]

ISTRUZIONI_EN = [
    ("Inventory template - laptops and tablets", True),
    ("", False),
    ('Fill in the "Inventario" sheet and import it from the program with', False),
    ("Import xls...  The blue rows carrying a room name are separators:", False),
    ("everything you write below one of them goes into that room, until", False),
    ("the next separator.", False),
    ("", False),
    ("Rules", True),
    ("- The Asset Tag is the only thing required: without it the row is", False),
    ("  discarded. Type, room and notes can be left empty.", False),
    ("- The template has the same columns as an exported file: you can", False),
    ("  export, fix things in Excel and import back.", False),
    ("- Need other columns? Add them: model, serial number and status are", False),
    ("  recognised by name and imported all the same.", False),
    ("- The asset tag identifies the device: importing the same asset tag", False),
    ("  twice updates the record instead of duplicating it.", False),
    ("- Type and Room have dropdowns: use the values offered.", False),
    ("- You can leave the room empty: the separators declare it.", False),
    ("", False),
    ("Copy and paste", True),
    ("- From here you can copy whole columns and paste them into the", False),
    ("  program: in Delete + you paste asset tags, one per line.", False),
    ("- In the program every field accepts Ctrl+C and Ctrl+V, and the", False),
    ("  right mouse button opens Copy / Paste.", False),
    ("- From the program's list you copy the selected row with Ctrl+C,", False),
    ("  or just the identifier with the right button: it pastes in", False),
    ("  here as it is.", False),
    ("- You can add rows under a separator, or move the separators.", False),
    ("- Do not change the column names in the first row.", False),
    ("- Rows left empty are simply ignored.", False),
    ("", False),
    ("Allowed values", True),
    ("TIPI", False),
    ("STATI", False),
    ("The dropdowns work in Excel and LibreOffice. Numbers does not import", False),
    ("them: in that case type the values above, they are the same.", False),
    ("", False),
    ("iPhones are not imported", True),
    ("They are handled by hand in the program only, and appear neither in", False),
    ("imports nor in exports. Any you put here will be ignored.", False),
]


def build_template(path, rooms, stati=None, lingua=None):
    """Genera il modello vuoto da compilare e reimportare.

    Ha le stesse colonne di un file esportato, con le righe gia' divise per
    stanza dai separatori e le tendine sui campi a scelta fissa. Gli iPhone non
    compaiono: si inseriscono a mano dal programma.

    Le colonne che qui non ci sono - modello, numero di serie, stato - restano
    comunque importabili: chi ha un foglio suo che le contiene lo carica lo
    stesso, perche' l'importazione riconosce le colonne dal nome.

    lingua: "en" per intestazioni, tendine e istruzioni in inglese.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    stati = list(stati or STATI)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    intestazioni = [intestazione(HEADERS[f], lingua) for f in TEMPLATE_FIELDS]
    ws.append(intestazioni)
    for cella in ws[1]:
        cella.font = Font(bold=True, color="FFFFFF")
        cella.fill = _HEADER_FILL
        cella.alignment = Alignment(vertical="center")
        cella.border = _BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    fill_tag = PatternFill("solid", fgColor="DCE6F1")
    riga = 2
    for stanza in rooms:
        cella = ws.cell(row=riga, column=1, value=str(stanza).upper())
        cella.font = Font(bold=True, color="1F4E79")
        cella.fill = fill_tag
        ws.merge_cells(start_row=riga, start_column=1,
                       end_row=riga, end_column=len(TEMPLATE_FIELDS))
        ws.row_dimensions[riga].height = 18
        riga += 1
        for _ in range(RIGHE_PER_STANZA):
            for colonna in range(1, len(TEMPLATE_FIELDS) + 1):
                ws.cell(row=riga, column=colonna).border = _BORDER
            riga += 1

    ultima = riga - 1
    tipi = DataValidation(type="list", formula1='"%s"' % ",".join(TEMPLATE_TIPI),
                          allow_blank=True, showDropDown=False)
    if lingua == "en":
        tipi.error, tipi.errorTitle = "Choose Laptop or Tablet.", "Invalid type"
    else:
        tipi.error, tipi.errorTitle = "Scegli Laptop o Tablet.", "Tipo non valido"
    ws.add_data_validation(tipi)
    colonna_tipo = get_column_letter(TEMPLATE_FIELDS.index("tipo") + 1)
    tipi.add("%s2:%s%d" % (colonna_tipo, colonna_tipo, ultima))

    stati_mostrati = [traduci_stato(v, lingua) for v in stati]
    scelte = DataValidation(type="list", formula1='"%s"' % ",".join(stati_mostrati),
                            allow_blank=True, showDropDown=False)
    if lingua == "en":
        scelte.error, scelte.errorTitle = "Choose one of the listed statuses.", "Invalid status"
    else:
        scelte.error, scelte.errorTitle = "Scegli uno degli stati previsti.", "Stato non valido"
    if "stato" in TEMPLATE_FIELDS:
        ws.add_data_validation(scelte)
        colonna_stato = get_column_letter(TEMPLATE_FIELDS.index("stato") + 1)
        scelte.add("%s2:%s%d" % (colonna_stato, colonna_stato, ultima))

    # La stanza la dicono le righe separatore, ma la colonna c'e' comunque
    # perche' un file esportato ce l'ha: chi preferisce scriverla riga per riga
    # trova la tendina, chi si affida ai separatori la lascia vuota.
    if "stanza" in TEMPLATE_FIELDS and rooms:
        stanze = DataValidation(type="list",
                                formula1='"%s"' % ",".join(str(r) for r in rooms),
                                allow_blank=True, showDropDown=False)
        if lingua == "en":
            stanze.error = "Choose one of the configured rooms."
            stanze.errorTitle = "Invalid room"
        else:
            stanze.error = "Scegli una delle stanze configurate."
            stanze.errorTitle = "Stanza non valida"
        ws.add_data_validation(stanze)
        colonna_stanza = get_column_letter(TEMPLATE_FIELDS.index("stanza") + 1)
        stanze.add("%s2:%s%d" % (colonna_stanza, colonna_stanza, ultima))

    for i, campo in enumerate(TEMPLATE_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larghezza_colonna(
            intestazione(HEADERS[campo], lingua), [TEMPLATE_ESEMPI[campo]])

    guida = wb.create_sheet("Istruzioni" if lingua != "en" else "Instructions")
    testo = list(ISTRUZIONI_EN if lingua == "en" else ISTRUZIONI_IT)
    etichette = {"TIPI": ("Type:  " if lingua == "en" else "Tipo:  ") + ",  ".join(TEMPLATE_TIPI),
                 "STATI": ("Status: " if lingua == "en" else "Stato: ")
                          + ",  ".join(stati_mostrati)}
    for numero, (frase, grassetto) in enumerate(testo, start=1):
        cella = guida.cell(row=numero, column=1, value=etichette.get(frase, frase))
        if grassetto:
            cella.font = Font(bold=True, size=12, color="1F4E79")
    guida.column_dimensions["A"].width = 78

    try:
        wb.save(path)
    except Exception as exc:
        raise InventoryError("Impossibile scrivere %s:\n%s" % (path, exc))
    finally:
        wb.close()
    return path


def _percorso_outlook():
    """Dove sta outlook.exe su questo computer, o None."""
    if not sys.platform.startswith("win"):
        return None
    try:
        import winreg
    except ImportError:
        return None
    chiavi = (
        (winreg.HKEY_LOCAL_MACHINE,
         r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\OUTLOOK.EXE"),
        (winreg.HKEY_CURRENT_USER,
         r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\OUTLOOK.EXE"),
    )
    for radice, percorso in chiavi:
        try:
            with winreg.OpenKey(radice, percorso) as chiave:
                valore = winreg.QueryValueEx(chiave, "")[0]
            if valore and os.path.exists(valore):
                return valore
        except OSError:
            continue
    # installazioni che non registrano App Paths
    for base in (os.environ.get("ProgramFiles", ""),
                 os.environ.get("ProgramFiles(x86)", "")):
        if not base:
            continue
        for ufficio in ("root\\Office16", "Office16", "Office15", "Office14"):
            candidato = os.path.join(base, "Microsoft Office", ufficio, "OUTLOOK.EXE")
            if os.path.exists(candidato):
                return candidato
    return None


def outlook_disponibile():
    return _percorso_outlook() is not None


def allega_a_outlook(percorsi):
    """Apre un nuovo messaggio di Outlook con i file gia' allegati.

    Outlook accetta un allegato solo dalla riga di comando: se i file sono piu'
    di uno vengono raccolti in un archivio zip, che e' anche piu' comodo da
    spedire. Ritorna il percorso di cio' che e' stato allegato.
    """
    percorsi = [p for p in percorsi if p and os.path.exists(p)]
    if not percorsi:
        raise InventoryError("Non c'e' nessun file da allegare.")
    outlook = _percorso_outlook()
    if outlook is None:
        raise InventoryError(
            "Outlook non e' stato trovato su questo computer.\n\n"
            "Il file e' stato creato lo stesso: allegalo a mano al messaggio.")
    if len(percorsi) == 1:
        allegato = percorsi[0]
    else:
        nome = "Inventario_%s.zip" % datetime.now().strftime("%Y%m%d_%H%M%S")
        allegato = os.path.join(os.path.dirname(percorsi[0]), nome)
        with zipfile.ZipFile(allegato, "w", zipfile.ZIP_DEFLATED) as archivio:
            for percorso in percorsi:
                archivio.write(percorso, os.path.basename(percorso))
    try:
        subprocess.Popen([outlook, "/a", allegato])
    except OSError as exc:
        raise InventoryError(
            "Non riesco ad aprire Outlook:\n%s\n\n"
            "Il file e' stato creato lo stesso: allegalo a mano." % exc)
    return allegato


def send_to_printer(path):
    """Invia il file alla stampante predefinita.

    Ritorna True se la stampa e' stata avviata, False se il file e' stato
    soltanto aperto (l'utente stampera' da Excel).
    """
    if sys.platform.startswith("win"):
        try:
            os.startfile(path, "print")  # noqa: attributo solo Windows
            return True
        except Exception:
            os.startfile(path)  # noqa
            return False
    opener = "open" if sys.platform == "darwin" else "xdg-open"
    try:
        subprocess.Popen([opener, path])
    except Exception as exc:
        raise InventoryError("Impossibile aprire il file di stampa:\n%s" % exc)
    return False


def open_file(path):
    if sys.platform.startswith("win"):
        os.startfile(path)  # noqa
        return
    opener = "open" if sys.platform == "darwin" else "xdg-open"
    subprocess.Popen([opener, path])
